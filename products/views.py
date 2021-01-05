from django.shortcuts import render
from .models import *

from testm.models import *
from testm.models import TestCase
from django.http import HttpResponse,JsonResponse
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required
import json
import re
# from testm.admin import ExportExcelMixin
from mysite.settings import OUTPUT_ROOT
from django.shortcuts import render
from docxtpl import DocxTemplate





# Create your views here.
# 产品联查指标、案例主函数
def products_dec(func1):

	def ProductsIndex(params0, params1, params2):
		class_ids = []
		records = []

		if params0 != {}:
			list1 = Products.objects.filter(id__in=params0['id'])
			class_id = []
			for var1 in list1:
				class_id.append(var1.prod_class_id)
			class_id = list(set(class_id))
			listA = ProductsClass.objects.filter(id__in=class_id)
			records.extend(func1(listA))
		else:
			if params1 != {}:
				list1 = Products.objects.filter(**params1)
				if params2 != {}:

					if 'location_type' in params2.keys():
						for var1 in list1:
							listA = ProductsClass.objects.filter(id=var1.prod_class_id)
							for var2 in listA:
								if var2.location_type == params2['location_type']:
									records.extend(func1(listA))
									break

					if 'performance_classification' in params2.keys():
						for var1 in list1:
							listA = ProductsClass.objects.filter(id=var1.prod_class_id)
							for var2 in listA:
								if var2.performance_classification == params2['performance_classification']:
									records.extend(func1(listA))
									break
					if 'prod_subclass' in params2.keys():
						for var1 in list1:
							listA = ProductsClass.objects.filter(id=var1.prod_class_id)
							for var2 in listA:
								if var2.prod_subclass == params2['prod_subclass']:
									records.extend(func1(listA))
									break
					if 'prod_type' in params2.keys():
						for var1 in list1:
							listA = ProductsClass.objects.filter(id=var1.prod_class_id)
							for var2 in listA:
								if var2.prod_type == params2['prod_type']:
									records.extend(func1(listA))
									break

				else:

					for var1 in list1:
						class_ids.append(var1.prod_class_id)
					listA = ProductsClass.objects.filter(id__in=class_ids)
					records.extend(func1(listA))

			else:
				if params2 != {}:
					kwargs = params2
					listA = ProductsClass.objects.filter(**kwargs)
					records.extend(func1(listA))
				else:
					list1 = Products.objects.all()
					class_ids = []
					for var1 in list1:
						class_ids.append(var1.prod_class_id)
					listA = ProductsClass.objects.filter(id__in=class_ids)
					records.extend(func1(listA))


		return records
	return ProductsIndex


# 类ID 查询指标函数(产品指标查询)
@products_dec
def ClassIndexget(listA):
	i = 0
	records_list = []
	if listA.exists():
		for varC in listA:
			listL = ProductsLocation.objects.filter(id=varC.location_type_id)
			listT = ProductsTypes.objects.filter(id=varC.prod_type_id)
			for varL,varT in zip(listL, listT):
				listI = Index.objects.filter(prod_class_id=varC.id)

				for varI in listI:
					listTW = TwoIndex.objects.filter(id=varI.two_index_id)
					for varTW in listTW:
						record_dic = dict()
						i = i + 1
						record_dic['test_type_name'] = str(varI.test_type)

						record_dic['location_type_name'] = varL.location_type_name
						record_dic['prod_type_name'] = varT.prod_type_name
						record_dic['prod_subclass'] = varC.prod_subclass
						record_dic['performance_classification'] = varC.performance_classification

						record_dic['first_index'] = str(varTW.first_index)
						record_dic['two_index'] = varTW.two_index
						record_dic['index_explanation'] = varTW.index_explanation
						record_dic['three_index'] = varI.index_name
						record_dic['index_description'] = varI.index_description
						record_dic['test_tools'] = varI.tool
						record_dic['test_steps'] = varI.test_steps
						record_dic['remarks'] = varI.remark

						if len(records_list) >=1:
							for j in range(0,len(records_list)):
								if records_list[j]['three_index'] == varI.index_name:
									break
								else:
									if j == len(records_list)-1:
										if str(record_dic['test_type_name']) == '1':
											record_dic['test_type_name'] = '验证'
										elif str(record_dic['test_type_name']) == '2':
											record_dic['test_type_name'] = '通过'
										elif str(record_dic['test_type_name']) == '3':
											record_dic['test_type_name'] = '评价'
										records_list.append(record_dic)
									else:
										continue
						else:
							if str(record_dic['test_type_name']) == '1':
								record_dic['test_type_name'] = '验证'
							elif str(record_dic['test_type_name']) == '2':
								record_dic['test_type_name'] = '通过'
							elif str(record_dic['test_type_name']) == '3':
								record_dic['test_type_name'] = '评价'
							records_list.append(record_dic)
	records_list.sort(key=lambda x: (x['test_type_name'], x['first_index'], x['two_index']))
	return records_list


# 类ID 查询测评指标函数
@products_dec
def ClassIndexTestget(listA):
	i = 0
	records_list = []
	if listA.exists():
		for varC in listA:
			listI = Index.objects.filter(prod_class_id=varC.id)

			for varI in listI:
				listTW = TwoIndex.objects.filter(id=varI.two_index_id)
				for varTW in listTW:
					record_dic = dict()
						# print(startNum,startNum+pageSize)
					i = i + 1

					record_dic['test_type_name'] = varI.test_type

					record_dic['first_index'] = str(varTW.first_index)
					record_dic['two_index'] = varTW.two_index
					record_dic['three_index'] = varI.index_name
					record_dic['test_score'] = varI.test_score
					record_dic['test_standard'] = varI.test_standard
					record_dic['score_description'] = varI.score_description

					if len(records_list) >=1:
						for j in range(0,len(records_list)):
							if records_list[j]['three_index'] == varI.index_name:
								break
							else:
								if j == len(records_list)-1:
									if str(record_dic['test_type_name']) == '1':
										record_dic['test_type_name'] = '验证'
									elif str(record_dic['test_type_name']) == '2':
										record_dic['test_type_name'] = '通过'
									elif str(record_dic['test_type_name']) == '3':
										record_dic['test_type_name'] = '评价'
									records_list.append(record_dic)
								else:
									continue
					else:
						if str(record_dic['test_type_name']) == '1':
							record_dic['test_type_name'] = '验证'
						elif str(record_dic['test_type_name']) == '2':
							record_dic['test_type_name'] = '通过'
						elif str(record_dic['test_type_name']) == '3':
							record_dic['test_type_name'] = '评价'
						records_list.append(record_dic)
	records_list.sort(key=lambda x: (x['test_type_name'], x['first_index'], x['two_index']))
	return records_list

# 类ID 查询案例函数
@products_dec
def ClassCaseget(listA):
	records_list = []
	i = 0
	if listA.exists():
		for var0 in listA:
			listI = Index.objects.filter(prod_class_id=var0.id)
			for varI in listI:
				listTW = TwoIndex.objects.filter(id=varI.two_index_id)

				for varTW in listTW:
					caselist = TestCase.objects.filter(case_name=varI.id)
					for varCase in caselist:
						record_dic = dict()
						# print(startNum,startNum+pageSize)
						i = i + 1
						record_dic['first_index'] = str(varTW.first_index)
						record_dic['two_index'] = varTW.two_index
						record_dic['case_name'] = str(varCase.case_name)
						record_dic['test_date'] = varCase.test_date
						record_dic['test_location'] = varCase.test_location
						record_dic['prod_name'] = varCase.prod_name
						record_dic['vend_name'] = varCase.vend_name
						record_dic['test_purpose'] = varCase.test_purpose
						record_dic['test_design'] = varCase.test_design
						record_dic['test_conditions'] = varCase.test_conditions
						record_dic['test_procedure'] = varCase.test_procedure
						record_dic['expected_results'] = varCase.expected_results
						record_dic['test_result'] = varCase.test_result
						record_dic['sign'] = varCase.sign
						record_dic['remark'] = varCase.remark
						if len(records_list) >= 1:
							for j in range(0, len(records_list)):
								if records_list[j]['case_name'] == varCase.case_name:
									break
								else:
									if j == len(records_list) - 1:
										records_list.append(record_dic)
									else:
										continue
						else:
							records_list.append(record_dic)

	records_list.sort(key=lambda x: (x['first_index'], x['two_index']))
	return records_list

# 导出excel函数功能
def ExportExcel(res,metas):
	wb = Workbook()
	ws = wb.active
	ws.append(metas['field_names'])
	response = HttpResponse(content_type='application/msexcel')
	#response = HttpResponse(content_type='application/json')
	response['Content-Disposition'] = 'attachment; filename='+metas['meta']+'.xlsx'
	if len(res) == 0:
		wb.save(response)
		return response
	for i in range(0, len(res)):
		data = [res[i][field] for field in metas['fields']]
		ws.append(data)
	wb.save(response)
	return response

# 导出excel函数(合并单元格)功能
def ExportExcelMerge(res, metas):
	wb = Workbook()
	ws = wb.active
	ws.append(metas['field_names'])
	response = HttpResponse(content_type='application/msexcel')
	response['Content-Disposition'] = 'attachment; filename='+metas['meta']+'.xlsx'
	if len(res) == 0:
		wb.save(response)
		return response
	for i in range(0, len(res)):
		data = [res[i][field] for field in metas['fields']]
		ws.append(data)
	# 获取第一列数据

	type_list1 = []
	type_list2 = []
	type_list3 = []
	i = 2
	for j in range(1,4):
		while True:
			r = ws.cell(i, j).value

			if r:
				if j == 1:
					type_list1.append(r)
				elif j == 2:
					type_list2.append(r)
				elif j == 3:
					type_list3.append(r)
			else:
				i = 2
				break

			i += 1
	# 第1列合并单元格
	s = 0
	e = 0
	flag = type_list1[0]

	for i in range(len(type_list1)):
		if type_list1[i] != flag:
			flag = type_list1[i]
			e = i - 1
			if e >= s:
				ws.merge_cells("A" + str(s + 2) + ":A" + str(e + 2))
				s = e + 1
		if i == len(type_list1) - 1:
			e = i
			ws.merge_cells("A" + str(s + 2) + ":A" + str(e + 2))
	# 第二列合并单元格
	flag = type_list1[0]
	flag2 = type_list2[0]
	s = 0
	e = 0
	if len(type_list1) == len(type_list2):
		for i in range(0,len(type_list2)):
			if type_list1[i] == flag:
				flag = type_list1[i]

				if type_list2[i] != flag2:
					flag2 = type_list2[i]
					e = i - 1
					if e >= s:
						ws.merge_cells("B" + str(s + 2) + ":B" + str(e + 2))
						s = e + 1
				else:
					flag2 = type_list2[i]
					continue

				if i == len(type_list2) - 1:
					e = i
					ws.merge_cells("B" + str(s + 2) + ":B" + str(e + 2))
			else:
				e = i - 1
				if e >= s:
					ws.merge_cells("B" + str(s + 2) + ":B" + str(e + 2))
					s = e + 1
				flag = type_list1[i]
				flag2 = type_list2[i]
				continue

	flag = type_list1[0]
	flag2 = type_list2[0]
	flag3 = type_list3[0]
	s = 0
	e = 0
	if len(type_list1) == len(type_list2) and len(type_list2) == len(type_list3):
		for i in range(0,len(type_list3)):
			if type_list2[i] == flag2 and type_list1[i] == flag:
				flag = type_list1[i]
				flag2 = type_list2[i]
				if type_list3[i] != flag3:
					flag3 = type_list3[i]
					e = i - 1
					if e >= s:
						ws.merge_cells("C" + str(s + 2) + ":C" + str(e + 2))
						s = e + 1
				else:
					flag3 = type_list3[i]
					continue
				if i == len(type_list3) - 1:
					e = i
					ws.merge_cells("C" + str(s + 2) + ":C" + str(e + 2))
			else:
				e = i - 1
				if e >= s:
					ws.merge_cells("C" + str(s + 2) + ":C" + str(e + 2))
					s = e + 1
				flag = type_list1[i]
				flag2 = type_list2[i]
				flag3 = type_list3[i]
				continue



	wb.save(response)
	#print(response['Content-Disposition'])

	return response




# 产品指标按钮功能
def ProdExportIndex(request):
	try:
		#导出 ID list
		if request.method == "GET":

			params_dic1 = dict()
			params_dic2 = dict()
			params_dic0 = dict()
			if request.GET.get("id_array") not in [None,'']:
				ids = request.GET.get("id_array").split(',')
				ids = list(map(int, ids))
				params_dic0['id'] = ids
			else:
				if request.GET.get('prod_name') not in ['null',''] :
					params_dic1['prod_name__icontains'] = request.GET.get('prod_name')
				if request.GET.get('vend_name') not in ['null',''] :
					params_dic1['vend_name__icontains'] = request.GET.get('vend_name')
				if request.GET.get('location_type') != 'null':
					params_dic2['location_type'] = request.GET.get('location_type')
				if request.GET.get('performance_classification') != 'null':
					params_dic2['performance_classification'] =request.GET.get('performance_classification')
				if request.GET.get('prod_subclass') != 'null':
					params_dic2['prod_subclass'] =request.GET.get('prod_subclass')
				if request.GET.get('prod_type') != 'null':
					params_dic2['prod_type'] =request.GET.get('prod_type')
			res = ClassIndexget(params_dic0,params_dic1,params_dic2)
			metas = dict()

			metas['meta'] = 'prodindex'
			metas['field_names'] = ['测试类型', '位置类型', '产品类型', '子类','性能分类','一级指标', '二级指标', '指标说明', '三级指标', '指标描述', '工具', '测试步骤', '备注']
			metas['fields'] = ['test_type_name', 'location_type_name', 'prod_type_name', 'prod_subclass','performance_classification', 'first_index',
					  'two_index', 'index_explanation', 'three_index', 'index_description', 'test_tools', 'test_steps',
					  'remarks']
			res1 = ExportExcel(res,metas)
			return res1

	except Exception as e:
		print(e)

# 测评指标函数
def ExportIndexTest(request):
	try:
		#导出 ID list
		if request.method == "GET":
			params_dic1 = dict()
			params_dic2 = dict()
			params_dic0 = dict()
			if request.GET.get("id_array") not in [None,'']:
				ids = request.GET.get("id_array").split(',')
				ids = list(map(int, ids))
				params_dic0['id'] = ids
			else:
				if request.GET.get('prod_name') not in ['null','']:
					params_dic1['prod_name__icontains'] = request.GET.get('prod_name')
				if request.GET.get('vend_name') not in ['null','']:
					params_dic1['vend_name__icontains'] = request.GET.get('vend_name')
				if request.GET.get('location_type') != 'null':
					params_dic2['location_type'] =request.GET.get('location_type')
				if request.GET.get('performance_classification') != 'null':
					params_dic2['performance_classification'] =request.GET.get('performance_classification')
				if request.GET.get('prod_subclass') != 'null':
					params_dic2['prod_subclass'] =request.GET.get('prod_subclass')
				if request.GET.get('prod_type') != 'null':
					params_dic2['prod_type'] =request.GET.get('prod_type')
			#res = ProductsIndex(params_dic1,params_dic2)
			res = ClassIndexTestget(params_dic0, params_dic1, params_dic2)
			metas = dict()

			metas['meta'] = 'prodindextest'
			metas['field_names'] = ['测试类型',  '一级指标', '二级指标',  '三级指标', '评测分值', '评测标准', '评分说明']
			# field_names = ['测试类型', '一级指标', '二级指标', '指标说明', '三级指标', '指标描述', '工具', '测试步骤', '备注']
			metas['fields'] = ['test_type_name', 'first_index','two_index',  'three_index', 'test_score','test_standard','score_description']
			res1 = ExportExcelMerge(res,metas)
			return res1

		elif request.method == "POST":
			ids = request.POST['id']
			print('testids')
			#print(ids)

	except Exception as e:
		print(e)
'''
def ProdExportCaseExcel(request):
	try:
		#导出 ID list
		if request.method == "GET":
			params_dic1 = dict()
			params_dic2 = dict()
			if request.GET.get('prod_name') != 'null':
				params_dic1['prod_name'] = request.GET.get('prod_name')
			if request.GET.get('vend_name') != 'null':
				params_dic1['vend_name'] = request.GET.get('vend_name')
			if request.GET.get('prod_class_id') != 'null':
				params_dic2['prod_class_id'] =request.GET.get('prod_class_id')
			res = ClassCaseget(params_dic1,params_dic2)
			metas = dict()
			metas['meta'] = 'prodcase'
			metas['field_names'] = ['一级指标', '二级指标','案例名称','测试时间', '测试地点', '产品名', '供应商','测试目的', '测试设计', '测试条件', '测试步骤', '预期结果','测试结果','签名','备注']
			metas['fields'] = ['first_index','two_index','case_name','test_date','test_location','prod_name','vend_name','test_purpose', 'test_design','test_conditions','test_procedure','expected_results','test_result','sign','remark']
			res1 = ExportExcel(res,metas)
			return res1

	except Exception as e:
		print(e)

'''
# 统计页面函数
@login_required
def GetStatistics(request):
	counts = dict()
	prod_local_counts = dict()
	prod_counts = Products.objects.count()
	prodclass_counts = ProductsClass.objects.count()
	prodtype_counts = ProductsTypes.objects.count()
	prodlocation_counts = ProductsLocation.objects.count()
	Index_counts = Index.objects.count()
	TwoIndex_counts = TwoIndex.objects.count()
	FirstIndex_counts = FirstIndex.objects.count()
	case_counts = TestCase.objects.count()
	listI = Index.objects.all()
	location_type = ['机房', '基础架构层-计算资源', '基础架构层-前端', '基础架构层-网络', '通用服务层-技术应用'
	, '通用服务层-安全', '应用支持', '应用集成', '数据部署', '业务应用']
	#产品位置分类计数
	for i in location_type:
		classids = []
		listL = ProductsLocation.objects.filter(location_type_name=i)
		for varL in listL:
			listC = ProductsClass.objects.filter(location_type=varL.id)
			for varC in listC:
				classids.append(varC.id)
		prod_local_counts[i] = Products.objects.filter(prod_class__in=classids).count()
	# 三级指标根据一级指标技术

	index13_count = dict()
	list1 = FirstIndex.objects.all()
	for var1 in list1:
		index13_count[var1.first_index] = Index.objects.filter(first_index=var1.id).count()

	index23_count = dict()
	list2 = TwoIndex.objects.all()
	for var2 in list2:
		index23_count[var2.two_index] = Index.objects.filter(first_index=var2.id).count()
	#print(index23_count)
	# 多少产品中有指标
	prodcls = []
	for varI in listI:
		if varI.prod_class not in prodcls:
			prodcls.append(varI.prod_class)
		else:
			continue

	index_prods = Products.objects.filter(prod_class__in=prodcls).count()


	counts['产品'] = prod_counts
	counts['三级指标'] = Index_counts
	counts['二级指标'] = TwoIndex_counts
	counts['一级指标'] = FirstIndex_counts
	counts['案例'] = case_counts
	counts['有指标的产品'] = index_prods
	counts['产品分类'] = prodclass_counts
	counts['产品类型'] = prodtype_counts
	counts['产品位置类别'] = prodlocation_counts
	#print(prod_local_counts)
	push_data = {'counts': json.dumps(counts), 'prod_local_counts': json.dumps(prod_local_counts),'index13_count':json.dumps(index13_count)}

	return render(request, 'statistics.html',push_data)


def ExportWord(case_list):
	filename = 'prodcase.docx'
	base_url = OUTPUT_ROOT
	asset_url = base_url + '/case_template.docx'
	tpl = DocxTemplate(asset_url)
	context = dict()
	context['case_list'] = case_list
	tpl.render(context)
	#tpl.save(base_url + filename)
	response = HttpResponse(content_type='application/msword')
	response['Content-Disposition'] = 'attachment;filename="{}"'.format(filename)
	tpl.save(response)
	return response

# 产品界面导出测试案例
def ProdExportCase(request):
	try:
		#导出 ID list
		if request.method == "GET":
			params_dic1 = dict()
			params_dic2 = dict()
			params_dic0 = dict()
			if request.GET.get("id_array") not in [None,'']:
				ids = request.GET.get("id_array").split(',')
				ids = list(map(int, ids))
				params_dic0['id'] = ids
			else:
				if request.GET.get('prod_name') not in ['null','']:
					params_dic1['prod_name'] = request.GET.get('prod_name')
				if request.GET.get('vend_name') not in ['null','']:
					params_dic1['vend_name'] = request.GET.get('vend_name')
				if request.GET.get('location_type') != 'null':
					params_dic2['location_type'] =request.GET.get('location_type')
				if request.GET.get('performance_classification') != 'null':
					params_dic2['performance_classification'] =request.GET.get('performance_classification')
				if request.GET.get('prod_subclass') != 'null':
					params_dic2['prod_subclass'] =request.GET.get('prod_subclass')
				if request.GET.get('prod_type') != 'null':
					params_dic2['prod_type'] =request.GET.get('prod_type')

			res = ClassCaseget(params_dic0,params_dic1,params_dic2)
			metas = dict()
			metas['meta'] = 'prodcase'
			metas['field_names'] = ['一级指标', '二级指标',  '案例名称','测试时间', '测试地点', '产品名', '供应商','测试目的', '测试设计', '测试条件', '测试步骤', '预期结果','测试结果','签名','备注']
			metas['fields'] = ['first_index','two_index','case_name','test_date','test_location','prod_name','vend_name','test_purpose', 'test_design','test_conditions','test_procedure','expected_results','test_result','sign','remark']
			res1 = ExportWord(res)
			return res1

	except Exception as e:
		print(e)


# 产品类别添加界面中得下来菜单联动
def choose_location(request):
	"""查询产品位置类型"""

	locations = ProductsLocation.objects.all()

	l_lists = [{"l_id": locations.id, "l_name": locations.location_type_name} for locations in locations]
	l_info = {"l_lists": l_lists}

	return JsonResponse(l_info, safe=False)

# 产品类别添加界面中得下来菜单联动
def choose_prodtypes(request):
	"""查询产品类型"""
	l_id = request.GET.get('l_id')
	prodtypes = ProductsTypes.objects.filter(location_type_id=l_id)
	t_lists = [{"t_id": prodtype.id, "t_name": prodtype.prod_type_name} for prodtype in prodtypes]
	t_info = {"t_lists": t_lists}
	return JsonResponse(t_info, safe=False)