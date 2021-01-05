# -*- coding:utf-8 -*-
import openpyxl
import os
import json
import datetime
import collections
from django.shortcuts import render
from testm.models import Projects
from xuanxing.models import Options
from django.contrib.auth.models import User
from products.models import ProductsClass, Products
from django.http import JsonResponse
from xuanxing.models import XuanXingRank
from xuanxing.models import ZhuangXiangRank
from xuanxing.models import ChushiRank
from xuanxing.models import AddScoreRank
from xuanxing.models import MinusScoreRank
from xuanxing.models import FinalScore
from xuanxing.models import ManageUser
from xuanxing.models import AnliDoc
from xuanxing.models import CaseSamllType
from xuanxing.models import CaseDocx

from tools.tools import xuanxing_date_all
from tools.tools import quarter_judge
from tools.tools import format_string
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path


def items_input(request):
    """
    事项录入
    :param request:
    :return:
    """
    # if request.method == "GET":
    #     return render(request, "score_kaohe/input.html")
    if request.method == "GET":
        login_user = ManageUser.objects.filter(id=int(request.session.get("user_id", "0"))).first()
        deal_user_list = [[pj_obj.deal_user.all(), pj_obj] for pj_obj in Projects.objects.all()]
        pj_list = [pj_obj for d_user, pj_obj in deal_user_list if login_user in d_user]
        # 取当前最新季度的
        item_all = []
        for pj_obj in pj_list:
            item_name = pj_obj.project_name
            vend_num = len([vend_temp.vend_name for vend_temp in pj_obj.vend_prod.all()])
            doc_name_list = [doc_obj.doc_name for doc_obj in AnliDoc.objects.filter(project_id=pj_obj)]
            anli_num = 0
            for anli_doc_name in doc_name_list:
                anli_doc_obj = AnliDoc.objects.filter(doc_name=anli_doc_name, project_id=pj_obj).first()
                case_type_list = [sm_obj.id for sm_obj in
                                  CaseSamllType.objects.filter(doc_name=anli_doc_obj, project=pj_obj)]
                all_case_list = [case_obj for case_obj in CaseDocx.objects.filter(case_type__in=case_type_list)]
                anli_num += len(all_case_list)
            meet_dic = {"0": "未知", "1": "选型测试(上会)", "2": "选型测试(不上会)", "3": "选型评估"}
            meet_status = meet_dic[str(pj_obj.meeting_status)]
            anli_coefficient = "%.2f" % (anli_num ** 0.5)  # 案例系数
            scale_dic = {"1": [2, 2.5], "2": [1, 1.5], "3": [0.7, 1.2]}
            # 规模系数
            scale_coefficient = scale_dic[str(pj_obj.meeting_status)][0] if vend_num <= 5 else scale_dic[str(pj_obj.meeting_status)][1]
            item_list = [item_name, vend_num, anli_num, meet_status, anli_coefficient, scale_coefficient]
            item_all.append(item_list)
        return render(request, "score_kaohe/item_record.html", {"item_all": item_all})


def huping(request):
    if request.method == "GET":
        hp_user = request.session.get("username", "")
        xuanxing_list = []
        for xx_obj in XuanXingRank.objects.filter(is_rank="1"):
            percent_dic = {}
            state_percent = xx_obj.stage_percent.split(",")
            fangan_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[0])] for ff in
                           xx_obj.fangan_percent.split(",")]
            hj_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[1])] for ff in
                       xx_obj.hj_percent.split(",")]
            exec_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[2])] for ff in
                         xx_obj.exec_percent.split(",")]
            all_list = fangan_list + hj_list + exec_list
            for per in all_list:
                if per[0] not in percent_dic:
                    percent_dic[per[0]] = per[1]
                else:
                    percent_dic[per[0]] += per[1]
            percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]
            is_huping = "1"
            if hp_user in xx_obj.fangan_percent or hp_user in xx_obj.exec_percent or hp_user in xx_obj.hj_percent:
                is_huping = "0"
            xx_list = [xx_obj.id, xx_obj.name, xx_obj.hj_score, xx_obj.anli_score, xx_obj.scale_score,
                       xx_obj.xuanxing_type, percent_list, is_huping]
            huping_list = ["", "", ""]
            if hp_user in xx_obj.huping:
                hp_all = xx_obj.huping.split(",")
                for hp in hp_all:
                    if hp_user in hp:
                        huping_list = hp.split("_")[1:]

            xx_list.append(huping_list)
            xuanxing_list.append(xx_list)
        shuxing_dic = {"1": "中心重点工作", "2": "处室重点工作", "3": "一般事项"}
        zhuanxiang_list = [
            [zx_obj.id, zx_obj.name, zx_obj.zhuanxiang_desc, shuxing_dic[zx_obj.shuxing], zx_obj.zhuanxiang_score,
             [zx.split("_") for zx in zx_obj.zhuanxiang_percent.split(",")],
             "0" if hp_user in zx_obj.zhuanxiang_percent else "1", zx_obj.huping] for
            zx_obj in
            ZhuangXiangRank.objects.filter(is_rank="1")]
        # 添加默认值
        zhuanxiang_str = ""
        for zx in zhuanxiang_list:
            if hp_user in zx[7]:
                for hp_item in zx[7].split(","):
                    if hp_user in hp_item:
                        zhuanxiang_str = hp_item.split("_")[1]
            zx.append(zhuanxiang_str)

        chushi_list = [[zx_obj.id, zx_obj.name, zx_obj.chushi_desc, zx_obj.chushi_score,
                        [zx.split("_") for zx in zx_obj.chushi_percent.split(",")],
                        "0" if hp_user in zx_obj.chushi_percent else "1", zx_obj.huping] for zx_obj in
                       ChushiRank.objects.filter(is_rank="1")]
        # 添加默认值
        chushi_str = ""
        for zx in chushi_list:
            if hp_user in zx[6]:
                for hp_item in zx[6].split(","):
                    if hp_user in hp_item:
                        chushi_str = hp_item.split("_")[1]
            zx.append(chushi_str)
        ad_list = [[zx_obj.id, zx_obj.name, zx_obj.add_desc, zx_obj.add_score,
                    [zx.split("_") for zx in zx_obj.add_percent.split(",")],
                    "0" if hp_user in zx_obj.add_percent else "1", zx_obj.huping] for zx_obj in
                   AddScoreRank.objects.filter(is_rank="1")]

        # 添加默认值
        add_str = ""
        for zx in ad_list:
            if hp_user in zx[6]:
                for hp_item in zx[6].split(","):
                    if hp_user in hp_item:
                        add_str = hp_item.split("_")[1]
            zx.append(add_str)

        mi_list = [[zx_obj.id, zx_obj.name, zx_obj.minus_desc, zx_obj.minus_score,
                    [zx.split("_") for zx in zx_obj.minus_percent.split(",")],
                    "0" if hp_user in zx_obj.minus_percent else "1", zx_obj.huping] for zx_obj in
                   MinusScoreRank.objects.filter(is_rank="1")]
        # 添加默认值
        mi_str = ""
        for zx in mi_list:
            if hp_user in zx[6]:
                for hp_item in zx[6].split(","):
                    if hp_user in hp_item:
                        mi_str = hp_item.split("_")[1]
            zx.append(mi_str)
        quarter_all = xuanxing_date_all()
        return render(request, "score_kaohe/huping.html",
                      {"xuanxing_list": xuanxing_list, "zhuanxiang_list": zhuanxiang_list, "chushi_list": chushi_list,
                       "ad_list": ad_list, "mi_list": mi_list, "quarter_all": quarter_all})


def lingdao(request):
    if request.method == "GET":
        xx_score_dic = {"戴路": 0, "张勇": 0, "倪海波": 0, "姜炜": 0, "王超": 0, "杨博": 0}
        xuanxing_list = []
        xx_sort_dic = {"戴路": 0, "张勇": 1, "倪海波": 2, "姜炜": 3, "王超": 4, "杨博": 5}
        # 选型测试
        for xx_obj in XuanXingRank.objects.filter(is_rank="1"):
            # 参与人的占比
            percent_dic = {}
            state_percent = xx_obj.stage_percent.split(",")
            fangan_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[0])] for ff in
                           xx_obj.fangan_percent.split(",")]
            hj_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[1])] for ff in
                       xx_obj.hj_percent.split(",")]
            exec_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[2])] for ff in
                         xx_obj.exec_percent.split(",")]
            all_list = fangan_list + hj_list + exec_list
            for per in all_list:
                if per[0] not in percent_dic:
                    percent_dic[per[0]] = per[1]
                else:
                    percent_dic[per[0]] += per[1]
            percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]
            # 每个人点评的分数
            hp_list = [["", "", "", "", "", "", ""], ["", "", "", "", "", "", ""], ["", "", "", "", "", "", ""]]
            for hp in xx_obj.huping.split(","):
                if hp != "":
                    xx_index = xx_sort_dic[hp.split("_")[0]]
                    h_list = hp.split("_")[1:]
                    i = 0
                    for h_score in h_list:
                        hp_list[i][xx_index] = h_score
                        i += 1
            # 总分
            final_score = "%.2f" % sum(
                [float("%.2f" % (sum([int(yy) for yy in y if yy]) / len([yy for yy in y if yy]))) for y in hp_list if
                 len([yy for yy in y if yy]) != 0])
            final_score = "%.2f" % ((float(final_score) * 100) * (float(xx_obj.hj_score) * 10) * (
                    float(xx_obj.scale_score) * 100) * (float(xx_obj.anli_score) * 100) / 10000000)
            hp_score = final_score
            if xx_obj.final_score:
                final_score = xx_obj.final_score

            # 平均分
            for hh in hp_list:
                hh[6] = "%.2f" % (sum([int(y) for y in hh if y]) / len([y for y in hh if y])) if len(
                    [y for y in hh if y]) else ""

            # 参与人
            join_user = ["0", "0", "0", "0", "0", "0"]
            all_join = xx_obj.fangan_percent + xx_obj.hj_percent + xx_obj.exec_percent
            for usr in xx_sort_dic:
                if usr in all_join:
                    join_user[xx_sort_dic[usr]] = "%.2f" % (
                            (float(percent_dic[usr] / 100) * 100) * (float(final_score) * 100) / 10000)
                    if join_user[xx_sort_dic[usr]]:
                        xx_score_dic[usr] += float(join_user[xx_sort_dic[usr]])
            # 赋值
            score_list = ["", "", ""]
            xx_score = [y for y in xx_obj.lingdao_input.split(",") if y]
            if xx_score:
                score_list = xx_score
            xx_list = [xx_obj.id, xx_obj.name, xx_obj.hj_score, xx_obj.anli_score, xx_obj.scale_score,
                       xx_obj.xuanxing_type, hp_list, final_score, join_user, percent_list, score_list, hp_score]
            xuanxing_list.append(xx_list)

        # 专项测试
        zhuanxiang_all = []
        shuxing_dic = {"1": "中心重点工作", "2": "处室重点工作", "3": "一般事项"}
        for zx_obj in ZhuangXiangRank.objects.filter(is_rank="1"):
            # 人员占比
            percent_dic = {}
            state_percent = zx_obj.zhuanxiang_percent.split(",")
            for per in state_percent:
                percent_dic[per.split("_")[0]] = per.split("_")[1]
            percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]

            # 各个人的评分情况
            hp_list = ["", "", "", "", "", ""]
            for hp in zx_obj.huping.split(","):
                if hp != "":
                    xx_index = xx_sort_dic[hp.split("_")[0]]
                    hp_list[xx_index] = hp.split("_")[1]

            # 总分
            final_score = "%.2f" % (sum([int(y) for y in hp_list if y]) / len([y for y in hp_list if y])) if len(
                [y for y in hp_list if y]) else ""

            hp_final = final_score
            final_score = zx_obj.final_score if zx_obj.final_score else final_score
            # 参与人
            join_user = ["0", "0", "0", "0", "0", "0"]
            for usr in xx_sort_dic:
                if usr in zx_obj.zhuanxiang_percent:
                    join_user[xx_sort_dic[usr]] = "%.2f" % (
                            (float(int(percent_dic[usr]) / 100) * 100) * (
                            float(final_score) * 100) / 10000) if final_score else ""
                    if join_user[xx_sort_dic[usr]]:
                        xx_score_dic[usr] += float(join_user[xx_sort_dic[usr]])
            # 领导的分
            xx_score = zx_obj.lingdao_input if zx_obj.lingdao_input else ""
            zx_list = [zx_obj.id, zx_obj.name, zx_obj.zhuanxiang_desc, hp_list, final_score, join_user,
                       zx_obj.zhuanxiang_score, shuxing_dic[zx_obj.shuxing], percent_list, xx_score, hp_final]
            zhuanxiang_all.append(zx_list)

        # 处室贡献
        chushi_all = []
        for zx_obj in ChushiRank.objects.filter(is_rank="1"):
            # 人员占比
            percent_dic = {}
            state_percent = zx_obj.chushi_percent.split(",")
            for per in state_percent:
                percent_dic[per.split("_")[0]] = per.split("_")[1]
            percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]
            # 各个人的评分情况
            hp_list = ["", "", "", "", "", ""]
            for hp in zx_obj.huping.split(","):
                if hp != "":
                    xx_index = xx_sort_dic[hp.split("_")[0]]
                    hp_list[xx_index] = hp.split("_")[1]

            # 总分
            final_score = "%.2f" % (sum([int(y) for y in hp_list if y]) / len([y for y in hp_list if y])) if len(
                [y for y in hp_list if y]) else ""
            hp_final = final_score
            final_score = zx_obj.final_score if zx_obj.final_score else final_score
            # 参与人
            join_user = ["0", "0", "0", "0", "0", "0"]
            for usr in xx_sort_dic:
                if usr in zx_obj.chushi_percent:
                    join_user[xx_sort_dic[usr]] = "%.2f" % (
                            (float(int(percent_dic[usr]) / 100) * 100) * (
                            float(final_score) * 100) / 10000) if final_score else ""
                    if join_user[xx_sort_dic[usr]]:
                        xx_score_dic[usr] += float(join_user[xx_sort_dic[usr]])
            # 领导的分
            xx_score = zx_obj.lingdao_input if zx_obj.lingdao_input else ""
            zx_list = [zx_obj.id, zx_obj.name, zx_obj.chushi_desc, hp_list, final_score, join_user, zx_obj.chushi_score,
                       percent_list, xx_score, hp_final]
            chushi_all.append(zx_list)
        # 加分项
        add_all = []
        for zx_obj in AddScoreRank.objects.filter(is_rank="1"):
            # 人员占比
            percent_dic = {}
            state_percent = zx_obj.add_percent.split(",")
            for per in state_percent:
                percent_dic[per.split("_")[0]] = per.split("_")[1]
            percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]
            # 各个人的评分情况
            hp_list = ["", "", "", "", "", ""]
            for hp in zx_obj.huping.split(","):
                if hp != "":
                    xx_index = xx_sort_dic[hp.split("_")[0]]
                    hp_list[xx_index] = hp.split("_")[1]

            # 总分
            final_score = "%.2f" % (sum([int(y) for y in hp_list if y]) / len([y for y in hp_list if y])) if len(
                [y for y in hp_list if y]) else ""
            hp_final = final_score
            final_score = zx_obj.final_score if zx_obj.final_score else final_score
            # 参与人
            join_user = ["0", "0", "0", "0", "0", "0"]
            for usr in xx_sort_dic:
                if usr in zx_obj.add_percent:
                    join_user[xx_sort_dic[usr]] = "%.2f" % (
                            (float(int(percent_dic[usr]) / 100) * 100) * (
                            float(final_score) * 100) / 10000) if final_score else ""
                    if join_user[xx_sort_dic[usr]]:
                        xx_score_dic[usr] += float(join_user[xx_sort_dic[usr]])
            # 领导的分
            xx_score = zx_obj.lingdao_input if zx_obj.lingdao_input else ""
            zx_list = [zx_obj.id, zx_obj.name, zx_obj.add_desc, hp_list, final_score, join_user, zx_obj.add_score,
                       percent_list, xx_score, hp_final]
            add_all.append(zx_list)
        # 扣分项
        minus_all = []
        for zx_obj in MinusScoreRank.objects.filter(is_rank="1"):
            # 人员占比
            percent_dic = {}
            state_percent = zx_obj.minus_percent.split(",")
            for per in state_percent:
                percent_dic[per.split("_")[0]] = per.split("_")[1]
            percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]

            # 各个人的评分情况
            hp_list = ["", "", "", "", "", ""]
            for hp in zx_obj.huping.split(","):
                if hp != "":
                    xx_index = xx_sort_dic[hp.split("_")[0]]
                    hp_list[xx_index] = hp.split("_")[1]

            # 总分
            final_score = "%.2f" % (sum([int(y) for y in hp_list if y]) / len([y for y in hp_list if y])) if len(
                [y for y in hp_list if y]) else ""
            hp_final = final_score
            final_score = zx_obj.final_score if zx_obj.final_score else final_score
            final_score = -abs(float(final_score))

            # 参与人
            join_user = ["0", "0", "0", "0", "0", "0"]
            for usr in xx_sort_dic:
                if usr in zx_obj.minus_percent:
                    join_user[xx_sort_dic[usr]] = "%.2f" % (
                            (float(int(percent_dic[usr]) / 100) * 100) * (
                            float(final_score) * 100) / 10000) if final_score else ""
                    if join_user[xx_sort_dic[usr]]:
                        xx_score_dic[usr] += -abs(float(join_user[xx_sort_dic[usr]]))
            # 领导的分
            xx_score = -abs(zx_obj.lingdao_input) if zx_obj.lingdao_input else ""
            zx_list = [zx_obj.id, zx_obj.name, zx_obj.minus_desc, hp_list, final_score, join_user, zx_obj.minus_score,
                       percent_list, xx_score, hp_final]
            minus_all.append(zx_list)
        fs_obj = FinalScore.objects.filter(id=1).first()
        if fs_obj and fs_obj.score:
            sum_score_list = fs_obj.score.split(",")
            print("库里的")
        else:
            sum_score_list = [0, 0, 0, 0, 0, 0]
            for sc in xx_score_dic:
                sum_score_list[xx_sort_dic[sc]] = "%.2f" % xx_score_dic[sc]

        quarter_all = xuanxing_date_all()
        return render(request, "score_kaohe/lingdao.html",
                      {"xuanxing_list": xuanxing_list, "zhuanxiang_list": zhuanxiang_all, "chushi_list": chushi_all,
                       "add_list": add_all, "minus_list": minus_all, "sum_score_list": sum_score_list,
                       "quarter_all": quarter_all})


def item_submit(request):
    recv_data = request.POST.dict()  # 所有参数
    now_date = datetime.datetime.now().strftime("%Y-%m-%d")
    print(now_date)
    print(recv_data)
    xuanxing_all = recv_data.get("xuanxing_all", "[]")
    xuanxing_all = json.loads(xuanxing_all) if xuanxing_all else xuanxing_all

    zhuanxiang_all = recv_data.get("zhuanxiang_all", "[]")
    zhuanxiang_all = json.loads(zhuanxiang_all) if zhuanxiang_all else zhuanxiang_all

    chushi_all = recv_data.get("chushi_all", "[]")
    chushi_all = json.loads(chushi_all) if chushi_all else chushi_all

    add_score_all = recv_data.get("add_score_all", "[]")
    add_score_all = json.loads(add_score_all) if add_score_all else add_score_all

    minus_score_all = recv_data.get("minus_score_all", "[]")
    minus_score_all = json.loads(minus_score_all) if minus_score_all else minus_score_all

    xx_name_list = [xx.name for xx in XuanXingRank.objects.all()]
    zx_name_list = [zx.name for zx in ZhuangXiangRank.objects.all()]
    cs_name_list = [cs.name for cs in ChushiRank.objects.all()]
    ad_name_list = [ad.name for ad in AddScoreRank.objects.all()]
    mi_name_list = [mi.name for mi in MinusScoreRank.objects.all()]
    if xuanxing_all:
        for xx_item in xuanxing_all:
            if xx_item[0] in xx_name_list:
                return JsonResponse({"code": "0004"})
            else:
                xx_obj = XuanXingRank()
                xx_obj.date = now_date
                xx_obj.name = xx_item[0]
                xx_obj.changshang_num = xx_item[1]
                xx_obj.anli_num = xx_item[2]
                xx_obj.hj_level = xx_item[3]
                xx_obj.xuanxing_type = xx_item[4]
                xx_obj.hj_score = xx_item[5]
                xx_obj.anli_score = xx_item[6]
                xx_obj.scale_score = xx_item[7]
                xx_obj.stage_percent = ",".join([xx_item[8][0][1], xx_item[8][1][1], xx_item[8][2][1]])
                xx_obj.fangan_percent = ",".join(xx_item[8][0][2])
                xx_obj.hj_percent = ",".join(xx_item[8][1][2])
                xx_obj.exec_percent = ",".join(xx_item[8][2][2])
                jizhun_score = 30 if str(xx_item[4]) == "1" else 20
                xx_obj.xuanxing_score = "%.2f" % (
                        jizhun_score * (float(xx_item[5]) * 10) * (float(xx_item[6]) * 100) * (
                        float(xx_item[7]) * 10) / 10000)
                xx_obj.save()
    if zhuanxiang_all:
        for zx_item in zhuanxiang_all:
            if zx_item[0] in zx_name_list:
                return JsonResponse({"code": "0004"})
            else:
                zx_obj = ZhuangXiangRank()
                zx_obj.date = now_date
                zx_obj.name = zx_item[0]
                zx_obj.shuxing = zx_item[1]
                zx_obj.zhuanxiang_percent = ",".join(zx_item[2])
                zx_obj.zhuanxiang_desc = zx_item[3]
                zx_obj.zhuanxiang_score = abs(int(zx_item[4]))
                zx_obj.save()
    if chushi_all:
        for cs_item in chushi_all:
            if cs_item[0] in cs_name_list:
                return JsonResponse({"code": "0004"})
            else:
                cs_obj = ChushiRank()
                cs_obj.date = now_date
                cs_obj.name = cs_item[0]
                cs_obj.chushi_percent = ",".join(cs_item[1])
                cs_obj.chushi_desc = cs_item[2]
                cs_obj.chushi_score = abs(int(cs_item[3]))
                cs_obj.save()
    if add_score_all:
        for ad_item in add_score_all:
            if ad_item[0] in ad_name_list:
                return JsonResponse({"code": "0004"})
            else:
                ad_obj = AddScoreRank()
                ad_obj.date = now_date
                ad_obj.name = ad_item[0]
                ad_obj.add_percent = ",".join(ad_item[1])
                ad_obj.add_desc = ad_item[2]
                ad_obj.add_score = abs(int(ad_item[3]))
                ad_obj.save()
    if minus_score_all:
        for mi_item in minus_score_all:
            if mi_item[0] in mi_name_list:
                return JsonResponse({"code": "0004"})
            else:
                mi_obj = MinusScoreRank()
                mi_obj.date = now_date
                mi_obj.name = mi_item[0]
                mi_obj.minus_percent = ",".join(mi_item[1])
                mi_obj.minus_desc = mi_item[2]
                mi_obj.minus_score = -abs(mi_item[3])
                mi_obj.save()

    return JsonResponse({"code": "0000"})


def item_check(request):
    xx_type_dic = {"1": "现场测试(上会)", "2": "现场测试(不上会)", "3": "选型评估"}
    check_list = [["0", "未审核"], ["1", "审核通过"], ["2", "审核不通过"]]
    shixiang_dic = {"1": "中心重点工作", "2": "处室重点工作", "3": "一般事项"}
    quarter_all = xuanxing_date_all()
    xuanxing_list = [
        [xx_obj.name, xx_obj.changshang_num, xx_obj.anli_num, xx_obj.hj_level, xx_type_dic[xx_obj.xuanxing_type],
         xx_obj.hj_score, xx_obj.anli_score, xx_obj.scale_score, xx_obj.xuanxing_score,
         xx_obj.stage_percent.split(",")[0], [ff.split("_") for ff in xx_obj.fangan_percent.split(",")],
         xx_obj.stage_percent.split(",")[1], [hj.split("_") for hj in xx_obj.hj_percent.split(",")],
         xx_obj.stage_percent.split(",")[2], [zx.split("_") for zx in xx_obj.exec_percent.split(",")],
         xx_obj.id, str(xx_obj.is_rank), xx_obj.markinfo] for xx_obj in XuanXingRank.objects.all()]

    zhuanxiang_list = [
        [zx_obj.name, shixiang_dic[zx_obj.shuxing], [zx.split("_") for zx in zx_obj.zhuanxiang_percent.split(",")],
         zx_obj.zhuanxiang_desc, zx_obj.zhuanxiang_score, zx_obj.is_rank, zx_obj.id, zx_obj.markinfo] for zx_obj in
        ZhuangXiangRank.objects.all()]

    chuhsi_list = [[zx_obj.name, [zx.split("_") for zx in zx_obj.chushi_percent.split(",")], zx_obj.chushi_desc,
                    zx_obj.chushi_score, zx_obj.is_rank, zx_obj.id, zx_obj.markinfo] for zx_obj in
                   ChushiRank.objects.all()]

    add_list = [[zx_obj.name, [zx.split("_") for zx in zx_obj.add_percent.split(",")], zx_obj.add_desc,
                 zx_obj.add_score, zx_obj.is_rank, zx_obj.id, zx_obj.markinfo] for zx_obj in AddScoreRank.objects.all()]
    minus_list = [[zx_obj.name, [zx.split("_") for zx in zx_obj.minus_percent.split(",")], zx_obj.minus_desc,
                   zx_obj.minus_score, zx_obj.is_rank, zx_obj.id, zx_obj.markinfo] for zx_obj in
                  MinusScoreRank.objects.all()]

    return render(request, "score_kaohe/shenhe.html",
                  {"check_list": check_list, "xuanxing_list": xuanxing_list, "zhuanxiang_list": zhuanxiang_list,
                   "chuhsi_list": chuhsi_list, "add_list": add_list, "minus_list": minus_list,
                   "quarter_all": quarter_all})


def shenhe_change(request):
    recv_data = request.POST.dict()
    type_name = recv_data.get("type_name")
    status = recv_data.get("change_status")
    change_id = recv_data.get("change_id")
    shenhe_markinfo = recv_data.get("markinfo", "")
    change_obj = {}
    if type_name == "xx":
        change_obj = XuanXingRank.objects.filter(id=int(change_id)).first()
    elif type_name == "zx":
        change_obj = ZhuangXiangRank.objects.filter(id=int(change_id)).first()
    elif type_name == "cs":
        change_obj = ChushiRank.objects.filter(id=int(change_id)).first()
    elif type_name == "ad":
        change_obj = AddScoreRank.objects.filter(id=int(change_id)).first()
    elif type_name == "mi":
        change_obj = MinusScoreRank.objects.filter(id=int(change_id)).first()
    if change_obj:
        change_obj.is_rank = str(status)
        change_obj.markinfo = shenhe_markinfo
        change_obj.save()
    return JsonResponse({"code": "0000"})


def huping_submit(request):
    recv_data = request.POST.dict()
    user_name = recv_data.get("username")
    print(recv_data)
    print(user_name)
    xuanxing_all = recv_data.get("xuanxing_all", "[]")
    xuanxing_all = json.loads(xuanxing_all) if xuanxing_all else xuanxing_all

    zhuanxiang_all = recv_data.get("zhuanxiang_all", "[]")
    zhuanxiang_all = json.loads(zhuanxiang_all) if zhuanxiang_all else zhuanxiang_all

    chushi_all = recv_data.get("chushi_all", "[]")
    chushi_all = json.loads(chushi_all) if chushi_all else chushi_all

    add_score_all = recv_data.get("add_all", "[]")
    add_score_all = json.loads(add_score_all) if add_score_all else add_score_all

    minus_score_all = recv_data.get("minus_all", "[]")
    minus_score_all = json.loads(minus_score_all) if minus_score_all else minus_score_all

    if xuanxing_all:
        for xx_item in xuanxing_all:
            xx_obj = XuanXingRank.objects.filter(id=int(xx_item[0])).first()
            hp_list = xx_obj.huping.split(",")
            x_list = xx_item[1:]
            x_list.insert(0, user_name)
            x_str = "_".join(x_list)
            print("----")
            print(x_str)

            if user_name not in xx_obj.huping:
                if xx_obj.huping == "":
                    hp_str = x_str
                else:
                    hp_list.append(x_str)
                    hp_str = ",".join(hp_list)
            else:
                new_list = []
                for hp_item in hp_list:
                    if user_name in hp_item:
                        new_list.append(x_str)
                    else:
                        new_list.append(hp_item)
                hp_str = ",".join(new_list)
            xx_obj.huping = hp_str
            xx_obj.save()

    if zhuanxiang_all:
        for zx_item in zhuanxiang_all:
            zx_obj = ZhuangXiangRank.objects.filter(id=int(zx_item[0])).first()
            hp_list = zx_obj.huping.split(",")
            x_list = [user_name, str(zx_item[1])]
            x_str = "_".join(x_list)

            if user_name not in zx_obj.huping:
                hp_list.append(x_str)
                hp_str = ",".join(hp_list)
            else:
                new_list = []
                for hp_item in hp_list:
                    if user_name in hp_item:
                        new_list.append(x_str)
                    else:
                        new_list.append(hp_item)
                hp_str = ",".join(new_list)
            zx_obj.huping = hp_str
            zx_obj.save()
    if chushi_all:
        for cs_item in chushi_all:
            cs_obj = ChushiRank.objects.filter(id=int(cs_item[0])).first()
            hp_list = cs_obj.huping.split(",")
            x_list = [user_name, str(cs_item[1])]
            x_str = "_".join(x_list)

            if user_name not in cs_obj.huping:
                hp_list.append(x_str)
                hp_str = ",".join(hp_list)
            else:
                new_list = []
                for hp_item in hp_list:
                    if user_name in hp_item:
                        new_list.append(x_str)
                    else:
                        new_list.append(hp_item)
                hp_str = ",".join(new_list)
            cs_obj.huping = hp_str
            cs_obj.save()

    if add_score_all:
        for ad_item in add_score_all:
            ad_obj = AddScoreRank.objects.filter(id=int(ad_item[0])).first()
            hp_list = ad_obj.huping.split(",")
            x_list = [user_name, str(ad_item[1])]
            x_str = "_".join(x_list)
            if user_name not in ad_obj.huping:
                hp_list.append(x_str)
                hp_str = ",".join(hp_list)
            else:
                new_list = []
                for hp_item in hp_list:
                    if user_name in hp_item:
                        new_list.append(x_str)
                    else:
                        if hp_item != "":
                            new_list.append(hp_item)
                hp_str = ",".join(new_list)
            ad_obj.huping = hp_str
            ad_obj.save()
    if minus_score_all:
        for mi_item in minus_score_all:
            mi_obj = MinusScoreRank.objects.filter(id=int(mi_item[0])).first()
            hp_list = mi_obj.huping.split(",")
            x_list = [user_name, str(-abs(int(mi_item[1])))]
            x_str = "_".join(x_list)

            if user_name not in mi_obj.huping:
                hp_list.append(x_str)
                hp_str = ",".join(hp_list)
            else:
                new_list = []
                for hp_item in hp_list:
                    if user_name in hp_item:
                        new_list.append(x_str)
                    else:
                        if hp_item != "":
                            new_list.append(hp_item)
                hp_str = ",".join(new_list)
            mi_obj.huping = hp_str
            mi_obj.save()
    return JsonResponse({"code": "0000"})


def lingdao_submit(request):
    recv_data = request.POST.dict()
    lingdao_type = recv_data.get("lingdao_type", "")
    lingdao_id = recv_data.get("lingdao_id", "")
    lingdao_list = recv_data.get("lingdao_list", "[]")
    lingdao_list = json.loads(lingdao_list) if lingdao_list else lingdao_list
    if lingdao_list:
        if lingdao_type == "xx":
            xx_obj = XuanXingRank.objects.filter(id=int(lingdao_id)).first()
            xx_obj.lingdao_input = ",".join(lingdao_list[0])
            xx_obj.final_score = lingdao_list[2]
            xx_obj.hp_score = lingdao_list[1]
            xx_obj.save()
            return JsonResponse({"code": "0000"})
        elif lingdao_type == "mi":
            mi_obj = MinusScoreRank.objects.filter(id=int(lingdao_id)).first()
            mi_obj.lingdao_input = lingdao_list[0]
            mi_obj.final_score = -abs(float(lingdao_list[2]))
            mi_obj.hp_score = -abs(float(lingdao_list[1]))
            mi_obj.save()
            return JsonResponse({"code": "0000"})

        else:
            if lingdao_type == "zx":
                zx_obj = ZhuangXiangRank.objects.filter(id=int(lingdao_id)).first()
            elif lingdao_type == "cs":
                zx_obj = ChushiRank.objects.filter(id=int(lingdao_id)).first()
            else:
                zx_obj = AddScoreRank.objects.filter(id=int(lingdao_id)).first()
            zx_obj.lingdao_input = lingdao_list[0]
            zx_obj.final_score = lingdao_list[2]
            zx_obj.hp_score = lingdao_list[1]
            zx_obj.save()
            return JsonResponse({"code": "0000"})

    return JsonResponse({"code": "0005"})


def save_final(request):
    recv_data = request.POST.dict()
    score_list = recv_data.get("score_list", "[]")
    score_list = json.loads(score_list) if score_list else score_list
    if score_list:
        fs_obj = FinalScore.objects.filter(id=1).first()
        if not fs_obj:
            fs_obj = FinalScore()
        fs_obj.score = ",".join(score_list)
        fs_obj.save()
        return JsonResponse({"code": "0000"})
    return JsonResponse({"code": "0005"})


def clear_final(request):
    flag = request.POST.get("flag", "0")
    if "1" == flag:
        fs_obj = FinalScore.objects.filter(id=1).first()
        if fs_obj:
            fs_obj.score = ""
            fs_obj.save()
        return JsonResponse({"code": "0000"})
    return JsonResponse({"code": "0005"})


def shenhe_list(request):
    check_user = request.session.get("username", "")
    check_date = format_string(request.GET.get("date", ""))
    xx_all, zx_all, cs_all, ad_all, mi_all = [], [], [], [], []

    xuanxing_all = XuanXingRank.objects.filter(
        Q(fangan_percent__contains=check_user) | Q(hj_percent__contains=check_user) | Q(
            exec_percent__contains=check_user))
    if xuanxing_all and check_date:
        xuanxing_all = [xx_obj for xx_obj in xuanxing_all if quarter_judge(xx_obj.date) == check_date]
    if xuanxing_all:
        xx_all = [[xx_obj.name, xx_obj.final_score, "%.2f" % ((float(xx_obj.final_score) * 100) * sum(
            [int(fa.split("_")[1]) * int(xx_obj.stage_percent.split(",")[0]) for fa in xx_obj.fangan_percent.split(",")
             if check_user in fa] + [int(fa.split("_")[1]) * int(xx_obj.stage_percent.split(",")[1]) for fa in
                                     xx_obj.hj_percent.split(",") if check_user in fa] + [
                int(fa.split("_")[1]) * int(xx_obj.stage_percent.split(",")[2]) for fa in xx_obj.exec_percent.split(",")
                if check_user in fa]) / 1000000)] for xx_obj in xuanxing_all]
    zhuanxiang_all = ZhuangXiangRank.objects.filter(zhuanxiang_percent__contains=check_user)
    if zhuanxiang_all and check_date:
        zhuanxiang_all = [zx_obj for zx_obj in zhuanxiang_all if quarter_judge(zx_obj.date) == check_date]
    if zhuanxiang_all:
        zx_all = [[zx_obj.name, zx_obj.final_score, "%.2f" % ((float(zx_obj.final_score) * 100) * (sum(
            [int(zx_percent.split("_")[1]) for zx_percent in zx_obj.zhuanxiang_percent.split(",") if
             check_user in zx_percent])) / 10000)] for zx_obj in zhuanxiang_all]
    chushi_all = ChushiRank.objects.filter(chushi_percent__contains=check_user)
    if chushi_all and check_date:
        chushi_all = [cs_obj for cs_obj in chushi_all if quarter_judge(cs_obj.date) == check_date]
    if chushi_all:
        cs_all = [[cs_obj.name, cs_obj.final_score, "%.2f" % ((float(cs_obj.final_score) * 100) * (sum(
            [int(zx_percent.split("_")[1]) for zx_percent in cs_obj.chushi_percent.split(",") if
             check_user in zx_percent])) / 10000)] for cs_obj in chushi_all]
    add_all = AddScoreRank.objects.filter(add_percent__contains=check_user)
    if add_all and check_date:
        add_all = [ad_obj for ad_obj in add_all if quarter_judge(ad_obj.date) == check_date]
    if add_all:
        ad_all = [[ad_obj.name, ad_obj.final_score, "%.2f" % ((float(ad_obj.final_score) * 100) * (sum(
            [int(zx_percent.split("_")[1]) for zx_percent in ad_obj.add_percent.split(",") if
             check_user in zx_percent])) / 10000)] for ad_obj in add_all]
    minus_all = MinusScoreRank.objects.filter(minus_percent__contains=check_user)
    if minus_all and check_date:
        minus_all = [mi_obj for mi_obj in minus_all if quarter_judge(mi_obj.date) == check_date]
    if minus_all:
        mi_all = [[mi_obj.name, mi_obj.final_score, "%.2f" % ((float(mi_obj.final_score) * 100) * (sum(
            [int(zx_percent.split("_")[1]) for zx_percent in mi_obj.minus_percent.split(",") if
             check_user in zx_percent])) / 10000)] for mi_obj in minus_all]

    quarter_all = xuanxing_date_all()  # 所有事项包含的季度
    return render(request, "score_kaohe/shenhe_list.html",
                  {"xx_all": xx_all, "zx_all": zx_all, "cs_all": cs_all, "ad_all": ad_all, "mi_all": mi_all,
                   "quarter_all": quarter_all})


def kaohe_jili(request):
    quarter_all = xuanxing_date_all()  # 所有事项包含的季度
    members = ["戴路", "张勇", "倪海波", "姜炜", "王超", "杨博"]
    return render(request, "score_kaohe/kaohe_jili.html", {"quarter_all": quarter_all, "members": members})


def score_excel(request):
    wb = Workbook()
    sheet = wb["Sheet"]
    # 大标题
    sheet["A1"] = "测试中心科技产品选型处2020年第三季度人员考核表"
    sheet.merge_cells('A1:M1')  # 合并一个矩形区域中的单元格
    bold_itatic_16_font = Font(name='等线', size=16, italic=False, color=colors.BLACK, bold=True)
    sheet['A1'].font = bold_itatic_16_font
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # 小标题
    sheet["A2"] = "工作事项"
    sheet.merge_cells("A2:B3")  # todo
    sheet["C2"] = "完成情况"
    sheet.merge_cells("C2:C3")
    sheet["D2"] = "基础分值"
    sheet.merge_cells("D2:D3")
    sheet["E2"] = "各人得分(打分)情况"
    sheet.merge_cells("E2:J2")  # 合并
    sheet["K2"] = "领导打分"
    sheet.merge_cells("K2:K3")
    sheet["K2"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    sheet["K2"].font = Font(color=colors.BLACK, bold=True)
    sheet["L2"] = "备注"
    sheet.merge_cells("L2:L3")
    sheet["L2"].font = Font(color=colors.BLACK, bold=True)
    sheet["L2"].alignment = Alignment(horizontal='center', vertical='center')

    # 基本参数
    jizhun_row = 4
    all_types = ["选型测试", "专项工作", "处室贡献", "加分项", "扣分项"]

    # sx_sh_list = ["质量", "效率", "上会效果", "参与比例", "测试环境复杂系数", "测试案例系数", "测试规模系数(详见说明)", "得分"]
    sx_sh_list = ["质量", "效率", "上会效果", "得分"]
    # sx_list = ["质量", "效率", "参与比例", "测试环境复杂系数", "测试案例系数", "测试规模系数(详见说明)", "得分"]
    sx_list = ["质量", "效率", "得分"]
    member_all = [["戴路", "E"], ["张勇", "F"], ["姜炜", "G"], ["倪海波", "H"], ["杨博", "I"], ["王超", "J"]]
    # 把人员全部填上
    for mem in member_all:
        sheet[mem[1] + str(jizhun_row-1)] = mem[0]
    # 循环全部数据
    row_i = jizhun_row
    for xx_type in all_types:
        col_i = 2  # 记录开始的列
        xx_start_row = row_i
        sheet["A" + str(row_i)] = xx_type  # 先把大类放进表格
        sheet["A" + str(row_i)].font = Font(size=13, bold=True)
        sheet["A" + str(row_i)].alignment = Alignment(horizontal='center', vertical='center')
        # 选型测试的事项循环
        if xx_type == "选型测试":
            for xx_obj in XuanXingRank.objects.filter(is_rank="1"):
                # 计算每个人的个人得分
                # 参与人的占比
                percent_dic = {}
                state_percent = xx_obj.stage_percent.split(",")
                fangan_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[0])] for ff in
                               xx_obj.fangan_percent.split(",")]
                hj_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[1])] for ff in
                           xx_obj.hj_percent.split(",")]
                exec_list = [[ff.split("_")[0], int(ff.split("_")[1]) / 100 * int(state_percent[2])] for ff in
                             xx_obj.exec_percent.split(",")]
                all_list = fangan_list + hj_list + exec_list
                for per in all_list:
                    if per[0] not in percent_dic:
                        percent_dic[per[0]] = per[1]
                    else:
                        percent_dic[per[0]] += per[1]
                percent_list = [[percent, percent_dic[percent]] for percent in percent_dic]
                # 打分人的情况  {"张勇": [10,9]}
                mark_dic = {format_string(xx.split("_")[0]): xx.split("_")[1:] for xx in xx_obj.huping.split(",")}
                print(mark_dic)
                start_row = row_i  # 记录开始的行，以便后面合并单元格
                sheet[get_column_letter(col_i) + str(row_i)] = xx_obj.name  # 事项名添加到表格

                if xx_obj.xuanxing_type == "1":
                    xx_params = sx_sh_list
                else:
                    xx_params = sx_list

                for param in xx_params:
                    if param in ["质量", "效率", "上会效果"]:
                        sheet[get_column_letter(col_i + 1) + str(row_i)] = param
                        sheet[get_column_letter(col_i + 2) + str(row_i)] = "10"
                    # elif param == "测试环境复杂系数":
                    #     sheet[get_column_letter(col_i + 1) + str(row_i)] = param
                    #     sheet[get_column_letter(col_i + 2) + str(row_i)] = xx_obj.hj_score
                    # elif param == "测试案例系数":
                    #     sheet[get_column_letter(col_i + 1) + str(row_i)] = param
                    #     sheet[get_column_letter(col_i + 2) + str(row_i)] = xx_obj.anli_score
                    # elif param == "测试规模系数(详见说明)":
                    #     sheet[get_column_letter(col_i + 1) + str(row_i)] = param
                    #     sheet[get_column_letter(col_i + 2) + str(row_i)] = xx_obj.scale_score
                    elif param == "得分":
                        sheet[get_column_letter(col_i + 1) + str(row_i)] = param
                        sheet[get_column_letter(col_i + 1) + str(row_i)].font = Font(color=colors.BLACK, bold=True)
                        sheet[get_column_letter(col_i + 2) + str(row_i)] = xx_obj.final_score
                        sheet[get_column_letter(col_i + 2) + str(row_i)].font = Font(color=colors.BLACK, bold=True)

                        if xx_obj.final_score:
                            # 参与人的得分情况
                            for col_index in range(3,9):
                                mem_name = sheet[get_column_letter(col_i + col_index) + str(jizhun_row-1)].value
                                print(mem_name)
                                mem_name = format_string(mem_name)
                                # 得分人的得分情况
                                if mem_name in percent_dic:
                                    sheet[get_column_letter(col_i + col_index) + str(row_i)] = "%.2f" % ((float(percent_dic[mem_name]) * 100) * (100* float(xx_obj.final_score))/1000000)
                                    sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.RED)
                                else:
                                    # 打分人的打分情况
                                    score_i = 0
                                    if mem_name in mark_dic:
                                        for score in mark_dic[mem_name]:
                                            sheet[get_column_letter(col_i + col_index) + str(start_row + score_i)] = score
                                            sheet[get_column_letter(col_i + col_index) + str(start_row + score_i)].font = Font(
                                                size=12, italic=False, color=colors.BLUE)
                                            score_i += 1
                            #  领导打分
                            lingdao_score = xx_obj.lingdao_input.split(",")
                            lingdao_i = 0
                            for score in lingdao_score:
                                sheet[get_column_letter(col_i+9) + str(start_row + lingdao_i)] = score
                                lingdao_i += 1
                    row_i += 1
                xx_type_dic = {"1": "现场测试(上会)", "2": "现场测试(不上会)", "3": "选型评估"}
                # 备注栏
                mark_str = " 厂商数: %s  \n 测试案例数: %s\n 测试环境复杂程度: %s \n 选型类型: %s" % (xx_obj.changshang_num, xx_obj.anli_num, xx_obj.hj_score, xx_type_dic[str(xx_obj.xuanxing_type)])
                sheet[get_column_letter(col_i+10) + str(start_row)] = mark_str
                if start_row < row_i - 1:
                    sheet.merge_cells("B%s:B%s" % (start_row, row_i - 1))
                    sheet["B%s" % start_row].alignment = Alignment(horizontal='center', vertical='center')
                    sheet.merge_cells("L%s:L%s" % (start_row, row_i - 1))
                    sheet["L%s" % start_row].alignment = Alignment(wrapText=True)

        # 专项工作的事项循环
        if xx_type == "专项工作":
            for zx_obj in ZhuangXiangRank.objects.filter(is_rank="1"):
                start_row = row_i
                sheet[get_column_letter(col_i) + str(row_i)] = zx_obj.name
                sheet[get_column_letter(col_i + 1) + str(row_i)] = "得分"
                sheet[get_column_letter(col_i + 1) + str(row_i)].font = Font(color=colors.BLACK, bold=True)
                sheet[get_column_letter(col_i + 2) + str(row_i)] = zx_obj.final_score
                sheet[get_column_letter(col_i + 2) + str(row_i)].font = Font(color=colors.BLACK, bold=True)

                score_dic = {per.split("_")[0]: "%.2f" % (int(per.split("_")[1]) * (100 * float(zx_obj.final_score))/10000) for per in zx_obj.zhuanxiang_percent.split(",") if per} if zx_obj.final_score else {}
                huping_dic = {hp.split("_")[0]: hp.split("_")[1] for hp in zx_obj.huping.split(",") if hp} if zx_obj.final_score else {}
                # 每个人的得分或者打分情况
                if zx_obj.final_score:
                    # 参与人的得分情况
                    for col_index in range(3, 9):
                        mem_name = sheet[get_column_letter(col_i + col_index) + str(jizhun_row - 1)].value  # 拿到姓名
                        mem_name = format_string(mem_name)  # 格式化mem_name 即去除空格之类
                        # 得分和打分填进表格
                        if mem_name in score_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = score_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.RED)
                        if mem_name in huping_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = huping_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.BLUE)
                    sheet[get_column_letter(col_i+9) + str(row_i)] = zx_obj.lingdao_input  # 领导打分栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)] = zx_obj.zhuanxiang_desc  # 备注栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)].alignment = Alignment(wrapText=True)
                row_i += 1

                if start_row < row_i - 1:  # 该事项的起始行到结束行的行号对比
                    sheet.merge_cells("B%s:B%s" % (start_row, row_i - 1))
                    sheet["B%s" % start_row].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
        # 处室贡献
        if xx_type == "处室贡献":
            for cs_obj in ChushiRank.objects.filter(is_rank="1"):
                start_row = row_i
                sheet[get_column_letter(col_i) + str(row_i)] = cs_obj.name
                sheet[get_column_letter(col_i + 1) + str(row_i)] = "得分"
                sheet[get_column_letter(col_i + 1) + str(row_i)].font = Font(color=colors.BLACK, bold=True)
                sheet[get_column_letter(col_i + 2) + str(row_i)] = cs_obj.final_score
                sheet[get_column_letter(col_i + 2) + str(row_i)].font = Font(color=colors.BLACK, bold=True)

                score_dic = {per.split("_")[0]: "%.2f" % (int(per.split("_")[1]) * (100 * float(cs_obj.final_score))/10000) for per in cs_obj.chushi_percent.split(",") if per} if cs_obj.final_score else {}
                huping_dic = {hp.split("_")[0]: hp.split("_")[1] for hp in cs_obj.huping.split(",") if hp} if cs_obj.final_score else {}
                # 每个人的得分或者打分情况
                if cs_obj.final_score:
                    # 参与人的得分情况
                    for col_index in range(3, 9):
                        mem_name = sheet[get_column_letter(col_i + col_index) + str(jizhun_row - 1)].value  # 拿到姓名
                        mem_name = format_string(mem_name)  # 格式化mem_name 即去除空格之类
                        # 得分和打分填进表格
                        if mem_name in score_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = score_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.RED)
                        if mem_name in huping_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = huping_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.BLUE)
                    sheet[get_column_letter(col_i+9) + str(row_i)] = cs_obj.lingdao_input  # 打分栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)] = cs_obj.chushi_desc  # 备注栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)].alignment = Alignment(wrapText=True)

                row_i += 1
                if start_row < row_i - 1:
                    sheet.merge_cells("B%s:B%s" % (start_row, row_i - 1))
                    sheet["B%s" % start_row].alignment = Alignment(horizontal='center', vertical='center')
        # 加分项
        if xx_type == "加分项":
            for ad_obj in AddScoreRank.objects.filter(is_rank="1"):
                start_row = row_i
                sheet[get_column_letter(col_i) + str(row_i)] = ad_obj.name
                sheet[get_column_letter(col_i + 1) + str(row_i)] = "得分"
                sheet[get_column_letter(col_i + 1) + str(row_i)].font = Font(color=colors.BLACK, bold=True)
                sheet[get_column_letter(col_i + 2) + str(row_i)] = ad_obj.final_score
                sheet[get_column_letter(col_i + 2) + str(row_i)].font = Font(color=colors.BLACK, bold=True)

                score_dic = {per.split("_")[0]: "%.2f" % (int(per.split("_")[1]) * (100 * float(ad_obj.final_score))/10000) for per in ad_obj.add_percent.split(",") if per} if ad_obj.final_score else {}
                huping_dic = {hp.split("_")[0]: hp.split("_")[1] for hp in ad_obj.huping.split(",") if hp} if ad_obj.final_score else {}
                # 每个人的得分或者打分情况
                if ad_obj.final_score:
                    # 参与人的得分情况
                    for col_index in range(3, 9):
                        mem_name = sheet[get_column_letter(col_i + col_index) + str(jizhun_row - 1)].value  # 拿到姓名
                        mem_name = format_string(mem_name)  # 格式化mem_name 即去除空格之类
                        # 得分和打分填进表格
                        if mem_name in score_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = score_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.RED)
                        if mem_name in huping_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = huping_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.BLUE)
                    sheet[get_column_letter(col_i+9) + str(row_i)] = ad_obj.lingdao_input  # 领导打分栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)] = ad_obj.add_desc  # 备注栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)].alignment = Alignment(wrapText=True)
                row_i += 1
                if start_row < row_i - 1:
                    sheet.merge_cells("B%s:B%s" % (start_row, row_i - 1))
                    sheet["B%s" % start_row].alignment = Alignment(horizontal='center', vertical='center')
        # 减分项
        if xx_type == "扣分项":
            for mi_obj in MinusScoreRank.objects.filter(is_rank="1"):
                start_row = row_i
                sheet[get_column_letter(col_i) + str(row_i)] = mi_obj.name
                sheet[get_column_letter(col_i + 1) + str(row_i)] = "得分"
                sheet[get_column_letter(col_i + 1) + str(row_i)].font = Font(color=colors.BLACK, bold=True)
                sheet[get_column_letter(col_i + 2) + str(row_i)] = mi_obj.final_score
                sheet[get_column_letter(col_i + 2) + str(row_i)].font = Font(color=colors.BLACK, bold=True)
                # 得分打分
                score_dic = {per.split("_")[0]: "%.2f" % (int(per.split("_")[1]) * (100 * float(mi_obj.final_score))/10000) for per in mi_obj.minus_percent.split(",") if per} if mi_obj.final_score else {}
                huping_dic = {hp.split("_")[0]: hp.split("_")[1] for hp in mi_obj.huping.split(",") if hp} if mi_obj.final_score else {}
                # 每个人的得分或者打分情况
                if mi_obj.final_score:
                    # 参与人的得分情况
                    for col_index in range(3, 9):
                        mem_name = sheet[get_column_letter(col_i + col_index) + str(jizhun_row - 1)].value  # 拿到姓名
                        mem_name = format_string(mem_name)  # 格式化mem_name 即去除空格之类
                        # 得分和打分填进表格
                        if mem_name in score_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = score_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.RED)
                        if mem_name in huping_dic:
                            sheet[get_column_letter(col_i + col_index) + str(row_i)] = huping_dic[mem_name]
                            sheet[get_column_letter(col_i + col_index) + str(row_i)].font = Font(size=12, italic=False, color=colors.BLUE)
                    sheet[get_column_letter(col_i+9) + str(row_i)] = mi_obj.lingdao_input  # 领导打分栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)] = mi_obj.minus_desc  # 备注栏
                    sheet[get_column_letter(col_i + 10) + str(row_i)].alignment = Alignment(wrapText=True)
                row_i += 1
                if start_row < row_i - 1:
                    sheet.merge_cells("B%s:B%s" % (start_row, row_i - 1))
                    sheet["B%s" % start_row].alignment = Alignment(horizontal='center', vertical='center')
        if xx_start_row < row_i - 1:
            sheet.merge_cells("A%s:A%s" % (xx_start_row, row_i - 1))

    # 遍历所有行
    for i in range(1, sheet.max_row):
        sheet.row_dimensions[i].height = 18  # 调节每一行的高度
        sheet["D%s" % i].alignment = Alignment(horizontal='center', vertical='center')
        if i == 2:
            # 设置每个列第3行及之后的对齐方式
            align_coloum = ["A", "C", "D", "E"]
            for col in align_coloum:
                sheet[col + str(i)].font = Font(size=13, bold=True)
                sheet[col + str(i)].alignment = Alignment(horizontal='center', vertical='center')
        if i > 2:
            # 设置 E-J的对齐方式
            align_coloum = ["E", "F", "G", "H", "I", "J"]
            for align_col in align_coloum:
                sheet[align_col + str(i)].alignment = Alignment(horizontal='center', vertical='center')

    # 调节宽度
    width_dic = {"A": 30, "B": 40, "C": 15, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10, "I": 10, "J": 10, "K": 5, "L": 40}
    for k in width_dic:
        sheet.column_dimensions[k].width = width_dic[k]
        # 给每个人得分的那些列进行居中处理
        if "D" < k < "K":
            sheet[k + str(2)].alignment = Alignment(horizontal='center', vertical='center')
            sheet[k + str(3)].alignment = Alignment(horizontal='center', vertical='center')
    # 个性调整
    sheet.row_dimensions[2].height = 30  # 第三行的高度调整为30
    sheet.row_dimensions[1].height = 40  # 第二行的高度调整为40
    files_mulu = os.getcwd() + "/static/kaohe"
    wb.save(files_mulu + "/" + "年度评分表.xlsx")
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path("年度评分表.xlsx"))
    wb.save(response)
    return response
