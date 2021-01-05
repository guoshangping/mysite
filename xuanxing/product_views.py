# -*- coding:utf-8 -*-
import openpyxl
import os
import re
import time

from django.shortcuts import render
from django.http import JsonResponse
from django.http import HttpResponse
from xuanxing.models import ProductsType
from xuanxing.models import ProductsDetail
from xuanxing.models import ProductStructure
from xuanxing.models import FirstIndex
from xuanxing.models import TwoIndex
from xuanxing.models import Index
from testm.models import Projects
from django.db.models import Q
from django.utils.encoding import escape_uri_path

from tools.tools import save_files
from tools.tools import juge_user_pj
from tools.tools import format_string
from openpyxl import Workbook


# 上传产品目录文件
def up_prod_content(request):
    try:
        print('--file upload---')
        prod_file_mulu = os.getcwd() + "/static/upload_files/product_list"
        resp_file = request.FILES.get('file', "")
        print("===================upload up_prod_content")
        rsp_doc_name = resp_file.name
        repeat_filename = "bak" + rsp_doc_name
        print(rsp_doc_name)
        file_list = os.listdir(prod_file_mulu)
        if rsp_doc_name not in file_list and repeat_filename not in file_list:
            # 读取前端上传来的文件，并写入服务器目录下
            with open('static/upload_files/product_list/' + rsp_doc_name, 'wb') as f:
                for line in resp_file.chunks():
                    f.write(line)
            print("----over---")
            plist_save_db()
            return JsonResponse({"code": "0000"})
        else:
            return JsonResponse({"code": "0004"})
    except Exception as e:
        print("-------------file upload  exception----------%s" % e)
        return JsonResponse({"code": "0001"})


def plist_save_db():
    print("plist_save_db start--------")
    files_mulu = os.getcwd() + "/static/upload_files/product_list"
    if 1:
        file_list = os.listdir(files_mulu)
        struct_dic = dict()
        print(file_list)
        for file_name in file_list:
            if not file_name.startswith("bak"):
                show_mulu = files_mulu + "/" + file_name
                sheet_obj = openpyxl.load_workbook(show_mulu)  # 加载excel
                big_class_list = sheet_obj.sheetnames  # 大类的sheet
                for bg_class in big_class_list:
                    bg_class = format_string(bg_class)
                    print('--------%s' % bg_class)
                    struct_dic[bg_class] = {}  # 大类
                    bg_table = sheet_obj[bg_class]  # 大类表格
                    nrows = list(bg_table.rows)  # 每行数据

                    del nrows[0]  # 去掉头行
                    md_class_list = list(set([format_string(rw[0].value) for rw in nrows if rw[0].value]))  # 产品类型
                    # 遍历产品类型
                    for md_cls in md_class_list:

                        struct_dic[bg_class][md_cls] = {}
                        # 读取第二列(子类)
                        # 存入产品类型数据库表
                        s_prod_list = list(
                            set(['%s <span style="display:none;">' % format_string(sm_rw[1].value) +
                                 '%s|%s|%s' % (bg_class, md_cls, format_string(sm_rw[1].value)) + '</span>'
                                 for sm_rw in nrows if
                                 format_string(sm_rw[0].value) and format_string(sm_rw[0].value) == md_cls]))

                        # 将大中小类存入数据库
                        # 子类
                        s_prod_list_db = list(
                            set([format_string(sm_rw[1].value) for sm_rw in nrows if
                                 format_string(sm_rw[0].value) and format_string(sm_rw[0].value) == md_cls]))
                        # 循环子类
                        for s_prod in s_prod_list_db:
                            pd_strc_list = ProductStructure.objects.filter(product_structure_name=bg_class)  # 筛选大类名称
                            if not pd_strc_list:
                                pd_strc_obj = ProductStructure()
                                pd_strc_obj.product_structure_name = bg_class
                                pd_strc_obj.save()
                                pd_strc = pd_strc_obj
                            else:
                                pd_strc = pd_strc_list[0]

                            check_obj_list = ProductsType.objects.filter(child_type=s_prod)
                            if not check_obj_list:
                                # 存入产品目录历史表(A)
                                pd_type_obj = ProductsType()  # 产品类型表
                                pd_type_obj.struct_name = pd_strc
                                pd_type_obj.prod_type = md_cls
                                pd_type_obj.child_type = s_prod
                                pd_type_obj.save()

                                # 存入产品目录当前表(B)
                                pd_type_obj2 = ProductsType()
                                pd_type_obj2.struct_name = pd_strc
                                pd_type_obj2.prod_type = md_cls
                                pd_type_obj2.child_type = s_prod
                                pd_type_obj2.project_name = "B"
                                pd_type_obj2.save()

                        struct_dic[bg_class][md_cls] = s_prod_list  # 把小类放进字典 用以展示

                    try:
                        # 遍历每行数据
                        for pd_sw in nrows:
                            pd_sw_0 = format_string(pd_sw[0].value)  # 第0格的数据
                            pd_sw_1 = format_string(pd_sw[1].value)  # 第1格的数据
                            if pd_sw_0 and pd_sw_1:
                                pd_detail_obj = ProductsDetail()
                                pd_strc_obj = ProductStructure.objects.filter(product_structure_name=bg_class).first()
                                if not pd_strc_obj:
                                    continue
                                pdtype_obj = ProductsType.objects.filter(
                                    Q(struct_name=pd_strc_obj) & Q(prod_type=pd_sw_0) & Q(child_type=pd_sw_1))
                                if pdtype_obj:
                                    pd_detail_obj.product_type = pdtype_obj[0]
                                pd_detail_obj.vend_tactics = pd_sw[2].value
                                pd_detail_obj.lifecycle = pd_sw[3].value
                                pd_detail_obj.performance_type = pd_sw[4].value
                                pd_detail_obj.vend_name = pd_sw[5].value
                                pd_detail_obj.product_name = pd_sw[6].value
                                pd_detail_obj.main_index = pd_sw[7].value
                                pd_detail_obj.apply_scene = pd_sw[8].value
                                pd_detail_obj.work_scope = pd_sw[9].value
                                pd_detail_obj.work_years = pd_sw[10].value
                                pd_detail_obj.mark_info = pd_sw[11].value
                                pd_detail_obj.save()
                    except Exception as e:
                        print(e)
            os.rename(files_mulu + f"/{file_name}", files_mulu + "/bak" + file_name)


# 保存产品目录到数据库(单机版用)
def plist_save(request):
    print("plist_save start--------")
    files_mulu = os.getcwd() + "/static/upload_files/product_list"
    all_list = []

    if 1:
        file_list = os.listdir(files_mulu)
        print(file_list)
        struct_dic = dict()
        for file_name in file_list:
            show_mulu = files_mulu + "/" + file_name
            sheet_obj = openpyxl.load_workbook(show_mulu)  # 加载excel
            print("-------------+++++++")
            big_class_list = sheet_obj.sheetnames  # 大类的sheet
            for bg_class in big_class_list:
                struct_dic[bg_class] = {}  # 大类
                bg_table = sheet_obj[bg_class]  # 大类表格
                nrows = list(bg_table.rows)  # 每行数据
                del nrows[0]  # 去掉头行
                md_class_list = list(set([rw[0].value for rw in nrows if rw[0].value]))  # 产品类型
                # 遍历产品类型
                for md_cls in md_class_list:
                    struct_dic[bg_class][md_cls] = {}
                    # 读取第二列(子类)
                    s_prod_list = list(
                        set(['%s <span style="display:none;">' % sm_rw[1].value +
                             '%s|%s|%s' % (bg_class, md_cls, sm_rw[1].value) + '</span>'
                             for sm_rw in nrows if sm_rw[0].value and sm_rw[0].value == md_cls]))

                    struct_dic[bg_class][md_cls] = s_prod_list  # 把小类放进字典 用以展示

            for big_class in struct_dic:
                prod_dic = dict()
                prod_dic["title"] = big_class
                # 以下数据结构为layui的tree组件所需要的格式
                # "children": [{"title": prod if prod else "空"} for prod in
                #                                   struct_dic[big_class][middle_cls][small_cls]]
                prod_dic["children"] = [{"title": middle_cls if middle_cls else "空", "children": [
                    {"title": small_cls if small_cls else "空"} for
                    small_cls in struct_dic[big_class][middle_cls]]} for middle_cls in struct_dic[big_class]]
                # 放入列表，每一个prod_dic都是一个大类的所有信息
                all_list.append(prod_dic)
        return JsonResponse({"code": "0000", "prod_info": all_list})


# 取数据库的产品信息数据返给前端
def prod_show(request):
    all_list = []
    print("prod_show start--------")
    pj_name = request.POST.get("pj_name", "A")  # 项目A为不变的那个版本
    real_pj_name = pj_name
    pj_name = "A" if pj_name == "A" else "B"  # 项目名称只有A和B(其他名称全部转B)
    big_cls_name = ""
    pd_show_id = ""
    pd_show_md_cls = ""
    struct_dic = dict()
    if real_pj_name == "A" or real_pj_name == "B":
        big_class_list = [big_cls_obj.product_structure_name for big_cls_obj in ProductStructure.objects.all()]
    else:
        pj_obj = Projects.objects.filter(project_name=real_pj_name).first()
        big_class_list = [pj_obj.product_subclass.all()[0].struct_name.product_structure_name]
        pd_show_id = pj_obj.product_subclass.all().first().id
        pd_show_md_cls = pj_obj.product_subclass.all().first().prod_type
        big_cls_name = big_class_list[0]
    print("-----------big_class_list----%s" % big_class_list)
    for bg_class in big_class_list:
        struct_dic[bg_class] = {}  # 9大类
        bg_class_obj = ProductStructure.objects.filter(product_structure_name=bg_class).first()
        md_class_list = list(set([md_cls_obj.prod_type for md_cls_obj in ProductsType.objects.filter(
            Q(struct_name=bg_class_obj) & Q(project_name=pj_name))]))  # 产品类型
        for md_cls in md_class_list:
            struct_dic[bg_class][md_cls] = {}
            s_prod_list = list(set([(
                                        '<i style="color:#FFB800">' if child_obj.id == pd_show_id else '<i>') + '%s</i>' % child_obj.child_type + '<span style="display:none;">' + '%s|%s|%s' % (
                                        bg_class, md_cls, child_obj.child_type) + '</span>' for child_obj in
                                    ProductsType.objects.filter(
                                        Q(prod_type=md_cls) & Q(project_name=pj_name) & Q(struct_name=bg_class_obj))]))
            struct_dic[bg_class][md_cls] = s_prod_list  # 把小类放进字典 用以展示
    print("-------------------999")
    for big_class in struct_dic:
        prod_dic = dict()
        prod_dic["title"] = big_class
        # 以下数据结构为layui的tree组件所需要的格式
        prod_dic["children"] = [{
                                    "title": '<span style="color:#FFB800">%s</span>' % middle_cls if middle_cls == pd_show_md_cls else middle_cls,
                                    "children": [
                                        {"title": small_cls if small_cls else "空"} for
                                        small_cls in struct_dic[big_class][middle_cls]]} for middle_cls in
                                struct_dic[big_class]]
        # 放入列表，每一个prod_dic都是一个大类的所有信息
        all_list.append(prod_dic)
    return JsonResponse({"code": "0000", "prod_info": all_list, "big_cls_name": big_cls_name, "pd_show_id": pd_show_id})


# 产品表格展示(excel版本)
def prod_list(request):
    resp_dic = {}  # 返回给前端的数据
    recv_data = request.POST.dict()  # 接收传来的请求参数
    title = recv_data.get("title", "")  # layui里的tree节点的 tilte
    print("*******title%s" % title)
    p1 = re.compile(r'<span.*span>', re.S)  # 正则匹配模块
    match_content = re.findall(p1, title)[0]  # 用上面的匹配规则整个title
    big_cls, mid_cls, sm_cls = match_content[28:-7].split("|")  # 大类小类种类的列表
    # 读取产品目录表
    files_mulu = os.getcwd() + "/static/upload_files/product_list"  # 产品目录文件 所在的目录
    file_name = "产品目录.xlsx"  # 产品目录文件的 文件名
    show_mulu = files_mulu + "/" + file_name  # 产品目录文件的完整目录
    sheet_obj = openpyxl.load_workbook(show_mulu)  # 用openpyxl库 加载产品目录文件
    big_class_list = sheet_obj.sheetnames  # excel里的所有页签的名字(即大类的名字)
    # 列表
    all_list = []
    # 遍历excel的大类
    for bg_class in big_class_list:
        # excel里的大类与此次tree里点击的title的大类相匹配
        if bg_class == big_cls:
            bg_table = sheet_obj[bg_class]
            nrows = list(bg_table.rows)
            del nrows[0]
            for rw in nrows:
                # 产品类型匹配
                if mid_cls == rw[0].value:
                    # 产品子类匹配
                    if sm_cls == rw[1].value:
                        pdlist = [p_rw.value for p_rw in rw]
                        pdlist.append(big_cls)  # 用以后续查询使用
                        all_list.append(pdlist)
            resp_dic["code"] = "0000"
            resp_dic["prod_data"] = all_list
            print('--------------')
            print(resp_dic["prod_data"])
            return JsonResponse(resp_dic)


# 产品表格展示(数据库版本)
def prod_list_db(request):
    resp_dic = {}  # 返回给前端的数据
    recv_data = request.POST.dict()  # 接收传来的请求参数
    title = recv_data.get("title", "")  # layui里的tree节点的 tilte
    pj_name = recv_data.get("pj_name", "A")  # layui里的tree节点的 tilte
    real_pj_name = pj_name
    pj_name = "A" if pj_name == "A" else "B"
    print("*******title%s" % title)
    p1 = re.compile(r'<span.*span>', re.S)  # 正则匹配模块
    match_content = re.findall(p1, title)[0]  # 用上面的匹配规则整个title
    big_cls, mid_cls, sm_cls = match_content[28:-7].split("|")  # 大类中类小类
    big_cls, mid_cls, sm_cls = big_cls.strip(), mid_cls.strip(), sm_cls.strip()
    ps_obj = ProductStructure.objects.filter(product_structure_name=big_cls).first()
    print("*****&&&&&")
    print(big_cls, mid_cls, sm_cls)

    pdt_obj = Projects.objects.filter(project_name=real_pj_name).first().product_subclass.all().first()
    pdt_obj_ = ProductsType.objects.filter(
        Q(project_name=pj_name) & Q(struct_name=ps_obj) & Q(prod_type=mid_cls) & Q(child_type=sm_cls)).first()
    pdt_obj = pdt_obj if pdt_obj == pdt_obj_ else ""

    # 取项目名称
    project_name = ""
    for pj_obj in Projects.objects.all():
        if pdt_obj in pj_obj.product_subclass.all():
            project_name = pj_obj.project_name
            break

    if pdt_obj:
        all_list = [
            [mid_cls, sm_cls, pdd_obj.vend_name, pdd_obj.product_name, pdd_obj.main_index, pdd_obj.lifecycle,
             pdd_obj.performance_type,
             pdd_obj.vend_tactics, pdd_obj.work_scope, pdd_obj.apply_scene, pdd_obj.work_years,
             pdd_obj.mark_info, pdd_obj.flag, big_cls, pdd_obj.id, project_name] for pdd_obj in
            ProductsDetail.objects.filter(product_type=pdt_obj)]
        resp_dic["code"] = "0000"
        resp_dic["prod_data"] = all_list
        resp_dic["pd_type_id"] = pdt_obj.id
        resp_dic["proj_name"] = project_name
        print('--------------prod_data')
        print(resp_dic["prod_data"])
        return JsonResponse(resp_dic)
    else:
        return JsonResponse({"code": "0002"})


# 产品信息,点击子节点之后的返回信息
def prod_info(request):
    normal_cls = request.GET.get("normal_cls", "")
    obj_id = request.GET.get("obj_id", "")
    big_cls = request.GET.get("big_cls", "")
    normal_lsit = normal_cls.split(",")  # 分类
    return render(request, "xuanxing_html/prod_info.html",
                  {"product_info": normal_lsit, "obj_id": obj_id, "big_cls": big_cls})


def prod_type_modify(request):
    pj_name = request.GET.get("pj_name", "")
    bg_cls = request.GET.get("bg_cls", "")
    pdt_list = list(set([ptobj.prod_type for ptobj in ProductsType.objects.filter(project_name=pj_name)]))
    cdt_list = list(set([ptobj.child_type for ptobj in ProductsType.objects.filter(project_name=pj_name)]))
    return render(request, "xuanxing_html/prod_type_edit.html",
                  {"pj_name": pj_name, "pdt_list": pdt_list, "cdt_list": cdt_list, "bg_cls": bg_cls})


def prod_add(request):
    recv_data = request.POST.dict()
    pj_name = recv_data.get("pj_name", "")
    prod_type = recv_data.get("prod_type", "")
    prod_child_type = recv_data.get("prod_child_type", "")
    prod_strategy = recv_data.get("prod_strategy", "")
    life_status = recv_data.get("life_status", "")
    property_type = recv_data.get("property_type", "")
    vend_name = recv_data.get("vend_name", "")
    prod_name = recv_data.get("prod_name", "")
    main_index = recv_data.get("main_index", "")
    deploy_scope = recv_data.get("deploy_scope", "")
    apply_scene = recv_data.get("apply_scene", "")
    use_years = recv_data.get("use_years", "")
    markinfo = recv_data.get("markinfo", "")
    if not all([pj_name, prod_child_type, prod_type]):
        return JsonResponse({"code": "0002"})
    # 存储数据库
    pdt_obj = ProductsType.objects.filter(
        Q(project_name=pj_name) & Q(prod_type=str(prod_type)) & Q(child_type=str(prod_child_type))).first()
    if not pdt_obj:
        return JsonResponse({"code": "0001"})
    pd_obj = ProductsDetail()
    pd_obj.project_name = pj_name
    pd_obj.product_type = pdt_obj
    pd_obj.vend_tactics = prod_strategy
    pd_obj.lifecycle = life_status
    pd_obj.performance_type = property_type
    pd_obj.vend_name = vend_name
    pd_obj.product_name = prod_name
    pd_obj.main_index = main_index
    pd_obj.apply_scene = apply_scene
    pd_obj.work_scope = deploy_scope
    pd_obj.work_years = use_years
    pd_obj.mark_info = markinfo
    pd_obj.flag = "2"  # 新增
    pd_obj.save()
    print("---ppppppppppppppp----------")
    return JsonResponse({"code": "0000"})


def prod_save(request):
    recv_data = request.POST.dict()
    obj_id = recv_data.get("obj_id", "")

    pd_obj = ProductsDetail.objects.filter(id=int(obj_id)).first()
    pd_obj.vend_tactics = recv_data.get("vend_strategy", "")
    pd_obj.lifecycle = recv_data.get("lifecycle", "")
    pd_obj.performance_type = recv_data.get("performance_type", "")
    pd_obj.vend_name = recv_data.get("vend_name", "")
    pd_obj.product_name = recv_data.get("pd_name", "")
    pd_obj.main_index = recv_data.get("main_index", "")
    pd_obj.apply_scene = recv_data.get("apply_scene", "")
    pd_obj.work_scope = recv_data.get("work_scope", "")
    pd_obj.work_years = recv_data.get("work_years", "")
    pd_obj.mark_info = recv_data.get("mark_info", "")
    pd_obj.flag = "1"
    pd_obj.save()
    return JsonResponse({"code": "0000"})


def prod_type_add(request):
    pj_name = request.POST.get("pj_name", "B")
    prod_type_name = request.POST.get("prod_type_name", "")
    prod_cd_type_name = request.POST.get("cd_type", "")
    bg_cls = request.POST.get("bg_cls", "")
    ps_obj = ProductStructure.objects.filter(product_structure_name=bg_cls).first()
    if not ps_obj:
        return JsonResponse({"code": "0001"})
    pt_obj = ProductsType()
    pt_obj.project_name = pj_name
    pt_obj.struct_name = ps_obj
    pt_obj.prod_type = str(prod_type_name)
    pt_obj.child_type = str(prod_cd_type_name)
    pt_obj.save()
    return JsonResponse({"code": "0000"})


def prod_excel_save(pj_name, pj_real):
    if pj_real != "B":
        pd_type = Projects.objects.filter(project_name=pj_real).first().product_subclass.all().first()
        pd_list = ProductsDetail.objects.filter(project_name=pj_name, product_type=pd_type)
    else:
        pd_list = ProductsDetail.objects.filter(project_name=pj_name)
    prod_list = [[pd_obj.product_type.prod_type, pd_obj.product_type.child_type, pd_obj.vend_tactics, pd_obj.lifecycle,
                  pd_obj.performance_type, pd_obj.vend_name, pd_obj.product_name, pd_obj.main_index, pd_obj.apply_scene,
                  pd_obj.work_scope, pd_obj.work_years, pd_obj.mark_info] for pd_obj in pd_list]
    title_list = ["产品类型", "子类", "厂商策略", "生命周期", "性能分类", "厂商名称", "产品名称", "主要技术指标", "应用场景", "部署范围", "使用年限", "备注"]
    prod_list.insert(0, title_list)

    file_name = f"prod_{pj_name}" + ".xlsx"
    files_mulu = os.getcwd() + "/static/upload_files/prod_export"
    file_name_list = os.listdir(files_mulu)
    if file_name in file_name_list:
        os.remove(files_mulu + "/" + file_name)
    # 存储成excel
    wb = Workbook()
    wb.create_sheet("产品信息")
    ws = wb.get_sheet_by_name("产品信息")
    for sv in prod_list:
        ws.append(sv)

    ws = wb["Sheet"]
    wb.remove(ws)
    wb.save(files_mulu + "/" + file_name)


def prod_export(request):
    pj_name = request.GET.get("pj_name", "")
    pj_real = request.GET.get("pj_real", "")
    prod_excel_save(pj_name, pj_real)
    if pj_name:
        print(pj_name)
        file_name = f"prod_{pj_name}" + ".xlsx"
        files_mulu = os.getcwd() + "/static/upload_files/prod_export"
        file_list = os.listdir(files_mulu)
        if file_name in file_list:
            wb = openpyxl.load_workbook(files_mulu + "/" + file_name)
            response = HttpResponse(content_type='application/msexcel')
            response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(
                escape_uri_path(f"{pj_name}产品目录.xlsx"))
            wb.save(response)
            return response
        return HttpResponse("文件不存在")
    else:
        return HttpResponse("项目名不存在")


def prod_change(request):
    pj_name = request.POST.get("pj_name", "B")
    if request.POST.get("option", "") == "0":
        md_type_name = request.POST.get("pd_type_name", "")
        md_chg_name = request.POST.get("md_chg_name", "")
        for tp_obj in ProductsType.objects.filter(Q(project_name=pj_name) & Q(prod_type=md_type_name)):
            tp_obj.prod_type = md_chg_name
            tp_obj.save()
    else:
        cd_type_name = request.POST.get("cd_type_name", "")
        chg_cd_type = request.POST.get("chg_cd_type", "")
        chg_cd_obj = ProductsType.objects.filter(id=int(chg_cd_type)).first()
        if chg_cd_obj:
            chg_cd_obj.child_type = cd_type_name
            chg_cd_obj.save()
    return JsonResponse({"code": "0000"})


def vend_prod_add_html(request):
    if request.method == "GET":
        small_class = request.GET.get("small_class", "")
        return render(request, "xuanxing_html/vend_prod_add.html", {"small_class": small_class})
    else:
        small_class = request.POST.get("small_class", "")
        prod_name = request.POST.get("prod_name", "")
        vend_name = request.POST.get("vend_name", "")
        if not all([small_class, prod_name, vend_name]):
            return JsonResponse({"code": "0001"})
        pt_obj = ProductsType.objects.filter(id=int(small_class)).first()
        if not pt_obj:
            return JsonResponse({"code": "0002"})
        # 存入数据库
        vend_pd_obj = ProductsDetail()
        vend_pd_obj.project_name = "B"
        vend_pd_obj.product_type = pt_obj
        vend_pd_obj.vend_name = vend_name
        vend_pd_obj.product_name = prod_name
        vend_pd_obj.save()

        vend_pj = [pj_obj for pj_obj in Projects.objects.all() if pt_obj in pj_obj.product_subclass.all()]
        if vend_pj:
            vend_pj[0].vend_prod.add(vend_pd_obj)
            print("------==AAABBB=======")
        return JsonResponse({"code": "0000"})


def prod_type_change(request):
    if request.method == "GET":
        pj_name = request.GET.get("pj_name", "B")
        big_cls_list = [ps_obj.product_structure_name for ps_obj in ProductStructure.objects.all()]
        return render(request, "xuanxing_html/prod_type_change.html",
                      {"big_cls_list": big_cls_list})


def prod_del(request):
    del_obj_id = request.POST.get("del_obj_id", "")
    ProductsDetail.objects.filter(id=int(del_obj_id)).delete()
    return JsonResponse({"code": "0000"})


def product_a_list(request):
    return render(request, "xuanxing_html/tree_a.html")


# -------以下为指标相关的功能-----------


# 指标新增修改页面
def index_type_modify(request):
    pj_name = request.GET.get("pj_name", "")
    # pdt_pj_name = "A" if pj_name == "项目A" else "B"
    bg_cls = request.GET.get("bg_cls", "")
    # pdt_list = list(set([ptobj.prod_type for ptobj in ProductsType.objects.filter(project_name=pdt_pj_name)]))
    # cdt_list = list(set([ptobj.child_type for ptobj in ProductsType.objects.filter(project_name=pdt_pj_name)]))
    pj_obj = Projects.objects.filter(project_name=pj_name).first()
    pdt_list = [pj_obj.product_subclass.all()[0].prod_type]
    cdt_list = [pj_obj.product_subclass.all()[0].child_type]

    f_list = [f_idx_obj.first_index for f_idx_obj in FirstIndex.objects.all()]
    return render(request, "xuanxing_html/index_type_edit.html",
                  {"pj_name": pj_name, "pdt_list": pdt_list, "cdt_list": cdt_list, "bg_cls": bg_cls, "f_list": f_list,
                   })


# 指标目录
def index_list(request):
    uid = request.session.get("user_id")
    pj_list = juge_user_pj(uid)
    return render(request, "xuanxing_html/index_tree.html", {"pj_list": pj_list})


def index_a_list(request):
    return render(request, "xuanxing_html/index_a_tree.html")


# 指标文件存储
def up_index_content(request):
    # try:
    resp_file = request.FILES.get('file', "")
    file_name = resp_file.name
    index_mulu = os.getcwd() + "/static/upload_files/index_list/"
    print("===================upload up_index_content")
    repeat_filename = "bak" + file_name
    file_list = os.listdir(index_mulu)
    if file_name not in file_list and repeat_filename not in file_list:
        save_files(index_mulu, file_name, resp_file)
        # save_files(mulu, file_name, resp_file)
        # 读取前端上传来的文件，并写入服务器目录下
        print("----over---")
        index_save("项目A", file_name)
        return JsonResponse({"code": "0000"})
    else:
        return JsonResponse({"code": "0004"})


# 指标目录tree 展示（读取excel版）
def index_show(request):
    return JsonResponse({"code": "0001"})
    print("index_show start--------")
    files_mulu_prod = os.getcwd() + "/static/upload_files/product_list/"  # 产品目录文件
    files_mulu_index = os.getcwd() + "/static/upload_files/index_list/"
    file_name_prod = "产品目录.xlsx"
    file_name_index = "指标数据字典.xlsx"
    all_list = []
    start_t = time.time()
    # 一二级指标放入字典
    index_dic = dict()
    show_mulu_index = files_mulu_index + file_name_index  # 指标文件的目录
    sheet_obj = openpyxl.load_workbook(show_mulu_index)  # 加载指标数据excel
    index_table = sheet_obj["数据字典"]
    index_nrows = list(index_table.rows)  # 每行数据
    del index_nrows[0]  # 去掉头行
    index_1_list = list(set([rw[2].value for rw in index_nrows if rw[2].value]))  # 一级指标
    print("))))))))))))))))))))))")
    for idx_one in index_1_list:
        idx_two_list = list(set([row_indx[3].value for row_indx in index_nrows if row_indx[2].value == idx_one]))
        index_dic[idx_one] = idx_two_list
    print(index_dic)
    # 读取指标产品目录结构
    show_mulu_prod = files_mulu_prod + file_name_prod
    sheet_obj_prod = openpyxl.load_workbook(show_mulu_prod)  # 加载excel
    print("-------------+++++++")
    big_class_list = sheet_obj_prod.sheetnames  # 大类的sheet
    struct_dic_prod = dict()
    # 产品大中小类的字典结构实现
    for bg_class in big_class_list:
        struct_dic_prod[bg_class] = {}  # 大类
        bg_table = sheet_obj_prod[bg_class]  # 大类表格
        nrows = list(bg_table.rows)  # 每行数据
        del nrows[0]  # 去掉头行
        md_class_list = list(set([rw[0].value for rw in nrows if rw[0].value]))  # 产品类型
        # 遍历产品类型
        for md_cls in md_class_list:
            struct_dic_prod[bg_class][md_cls] = {}
            # 读取第二列(子类)
            s_prod_list = list(set([sm_rw[1].value for sm_rw in nrows if sm_rw[0].value and sm_rw[0].value == md_cls]))
            # 把子类放进字典
            for sm_cls in s_prod_list:
                # 把一二级指标放入结构字典中
                struct_dic_prod[bg_class][md_cls][sm_cls] = index_dic

    # tree展示数据
    for big_class in struct_dic_prod:
        prod_dic = dict()
        prod_dic["title"] = '<span style="color:blue;">%s</span>' % big_class
        # 以下数据结构为layui的tree组件所需要的格式
        # "children": [{"title": prod if prod else "空"} for prod in
        #                                   struct_dic[big_class][middle_cls][small_cls]]
        prod_dic["children"] = [{"title": middle_cls if middle_cls else "空", "children": [
            {"title": small_cls if small_cls else "空",
             "children": [
                 {"title": idx_one,
                  "children": [{"title": '<i style="color:blue;">%s</i> <span style="display:none;">' % idx_two +
                                         '%s|%s|%s|%s|%s' % (
                                             big_class, middle_cls, small_cls, idx_one,
                                             idx_two) + '</span>'} for idx_two in
                               struct_dic_prod[big_class][middle_cls][small_cls][
                                   idx_one]]} for idx_one in
                 struct_dic_prod[big_class][middle_cls][small_cls]]} for
            small_cls in struct_dic_prod[big_class][middle_cls]]} for middle_cls in struct_dic_prod[big_class]]
        # 放入列表，每一个prod_dic都是一个大类的所有信息
        all_list.append(prod_dic)
    end_t = time.time()
    print("耗时: %s s" % (end_t - start_t))
    return JsonResponse({"code": "0000", "prod_info": all_list})


# 指标目录tree展示 (读取数据库版)
def index_show_db(request):
    pj_name = request.POST.get("pj_name", "项目A")
    idx_pjname = pj_name
    pj_name = "A" if pj_name == "项目A" else "B"
    print("index_show_db start--------")
    all_list = []
    big_cls_name = ""
    start_t = time.time()
    # 一二级指标放入字典
    index_dic = dict()
    for f_idx_obj in FirstIndex.objects.all():
        idx_two_list = [t_obj.two_index for t_obj in TwoIndex.objects.filter(first_index=f_idx_obj)]
        index_dic[f_idx_obj.first_index] = idx_two_list
    print(index_dic)
    idx_md_cls = ""
    idx_sm_cls = ""
    if pj_name == "A":
        big_class_list = [big_cls_obj.product_structure_name for big_cls_obj in ProductStructure.objects.all()]
    else:
        idx_obj = Projects.objects.filter(project_name=idx_pjname).first()
        big_class_list = [idx_obj.product_subclass.all()[0].struct_name.product_structure_name]
        big_cls_name = big_class_list[0]
        idx_md_cls = idx_obj.product_subclass.all()[0].prod_type
        idx_sm_cls = idx_obj.product_subclass.all()[0].child_type

    struct_dic_prod = dict()
    # 产品大中小类的字典结构实现
    for bg_class in big_class_list:
        ps_obj = ProductStructure.objects.filter(product_structure_name=bg_class)[0]
        struct_dic_prod[bg_class] = {}  # 大类
        md_sm_class_list = list(set(["%s" % pt_obj.prod_type + "$" + "%s" % pt_obj.child_type for pt_obj in
                                     ProductsType.objects.filter(
                                         Q(struct_name=ps_obj) & Q(project_name=pj_name))]))  # 产品类型
        # 遍历产品类型
        md_class_list = list(set([mdsm_str.split("$")[0] for mdsm_str in md_sm_class_list]))
        for md_cls in md_class_list:
            struct_dic_prod[bg_class][md_cls] = {}
            # 读取第二列(子类)
            s_prod_list = list(
                set([mdsm_str.split("$")[1] for mdsm_str in md_sm_class_list if mdsm_str.split("$")[0] == md_cls]))
            # 把子类放进字典
            for sm_cls in s_prod_list:
                # 把一二级指标放入结构字典中
                struct_dic_prod[bg_class][md_cls][sm_cls] = index_dic

    # tree展示数据
    for big_class in struct_dic_prod:
        prod_dic = dict()
        prod_dic["title"] = '<span style="color:blue;">%s</span>' % big_class
        # 以下数据结构为layui的tree组件所需要的格式
        # "children": [{"title": prod if prod else "空"} for prod in
        #                                   struct_dic[big_class][middle_cls][small_cls]]
        prod_dic["children"] = [{"title": "<span style='color:blue'>%s</span>" % middle_cls if middle_cls == idx_md_cls else middle_cls, "children": [
            {"title": '<span style="color:blue">%s</span>'% small_cls if small_cls == idx_sm_cls else small_cls,
             "children": [
                 {"title": idx_one,
                  "children": [{"title": '<i style="color:red;">%s</i> <span style="display:none;">' % idx_two +
                                         '%s|%s|%s|%s|%s' % (
                                             big_class, middle_cls, small_cls, idx_one,
                                             idx_two) + '</span>'} for idx_two in
                               struct_dic_prod[big_class][middle_cls][small_cls][
                                   idx_one]]} for idx_one in
                 struct_dic_prod[big_class][middle_cls][small_cls]]} for
            small_cls in struct_dic_prod[big_class][middle_cls]]} for middle_cls in struct_dic_prod[big_class]]
        # 放入列表，每一个prod_dic都是一个大类的所有信息
        all_list.append(prod_dic)
    end_t = time.time()
    print("耗时: %s s" % (end_t - start_t))
    return JsonResponse({"code": "0000", "prod_info": all_list, "big_cls_name": big_cls_name})


# 指标点击之后的table展示
def index_info(request):
    resp_dic = {}  # 返回给前端的数据
    recv_data = request.POST.dict()  # 接收传来的请求参数
    title = recv_data.get("title", "")  # layui里的tree节点的 tilte
    pj_name = recv_data.get("pj_name", "项目A")
    pdt_pj_name = "A" if pj_name == "项目A" else "B"

    title = title.strip()
    print("*******title%s" % title)
    p1 = re.compile(r'<span.*span>', re.S)  # 正则匹配模块
    match_content = re.findall(p1, title)[0]  # 用上面的匹配规则整个title
    big_cls, mid_cls, sm_cls, idx_one, idx_two = match_content[28:-7].split("|")  # 大类小类种类的列表
    print(big_cls, mid_cls, sm_cls, idx_one, idx_two)
    # 读取产品目录表
    ps_obj = ProductStructure.objects.filter(product_structure_name=big_cls).first()
    f_index_obj = FirstIndex.objects.filter(first_index=idx_one).first()  # 一级指标
    t_index_obj = TwoIndex.objects.filter(Q(two_index=idx_two) & Q(first_index=f_index_obj)).first()  # 二级指标
    pdt_obj = ProductsType.objects.filter(
        Q(project_name=pdt_pj_name) & Q(struct_name=ps_obj) & Q(prod_type=mid_cls) & Q(child_type=sm_cls))[0]
    print('===============')
    print(pdt_obj.prod_type)

    all_list = [[big_cls, mid_cls, sm_cls, idx_one, idx_two, idx_obj.index_explain, idx_obj.index_name,
                 idx_obj.index_description, idx_obj.index_id, idx_obj.anli_id, idx_obj.test_type, idx_obj.tool,
                 idx_obj.remark, idx_obj.flag, idx_obj.id] for idx_obj in Index.objects.all() if
                idx_obj.prod_class == pdt_obj and idx_obj.project_name == pj_name and idx_obj.first_index == f_index_obj and idx_obj.two_index == t_index_obj]
    print("--------------****")
    resp_dic["prod_data"] = all_list
    resp_dic["code"] = "0000"
    return JsonResponse(resp_dic)


def index_detail(request):
    normal_cls = request.GET.get("normal_cls", "")
    normal_lsit = normal_cls.split(",")  # 分类
    idx_id = request.GET.get("obj_id", "")
    return render(request, "xuanxing_html/index_info.html", {"product_info": normal_lsit, "obj_id": idx_id})


def index_save(pj_name, file_name):
    print("进入index-save")
    map_prod_class = {"机房资源": "0-机房", "计算资源": "1-基础架构层-计算资源", "前端": "2-基础架构层-前端", "网络": "3-基础架构层-网络",
                      "技术应用": "4-通用服务层-技术应用", "安全": "5通用服务层-安全", "应用支持": "6应用支持", "应用集成": "7应用集成",
                      "数据部署": "8数据部署", "业务应用": "9业务应用"}

    # 存储指标的字典
    index_dic = dict()
    # 文件目录
    file_mulu = os.getcwd() + "/static/upload_files/index_list/" + file_name

    sheet_obj = openpyxl.load_workbook(file_mulu)  # 加载指标数据excel
    # 指标类型
    index_table = sheet_obj["数据字典"]
    index_nrows = list(index_table.rows)  # 每行数据
    del index_nrows[0]  # 去掉头行

    # 具体的三级指标
    three_index_table = sheet_obj["计算资源产品"]
    three_index_nrows = list(three_index_table.rows)  # 每行数据
    del three_index_nrows[0]  # 去掉头行

    index_1_list = list(set([format_string(rw[5].value) for rw in index_nrows if rw[5].value]))  # 一级指标
    print("))))))))))))))))))))))")
    for idx_one in index_1_list:
        idx_two_list = list(set(
            [(format_string(row_indx[6].value), format_string(row_indx[7].value).zfill(4)) for row_indx in index_nrows
             if format_string(row_indx[5].value) == idx_one]))
        index_dic[idx_one] = idx_two_list
    # 将一二级指标存入数据库
    for index_k, index_v in index_dic.items():
        first_index_obj = FirstIndex.objects.filter(first_index=index_k).first()
        # 一级指标
        if not first_index_obj:
            first_obj = FirstIndex()
            first_obj.first_index = index_k
            first_obj.save()
            first_index_obj = first_obj

        # 二级指标的存储
        for t_index in index_v:
            print(t_index)
            two_index_obj = TwoIndex.objects.filter(Q(first_index=first_index_obj) & Q(two_index=t_index)).first()
            if not two_index_obj:
                two_obj = TwoIndex()
                two_obj.first_index = first_index_obj
                two_obj.two_index = t_index[0]
                two_obj.index_number = t_index[1]
                two_obj.save()
                two_index_obj = two_obj

            # 三级指标的存储
            for th_row in three_index_nrows:
                if format_string(th_row[3].value) == index_k and format_string(th_row[4].value) == t_index[0]:
                    print(th_row[0].value)
                    bg_class = ProductStructure.objects.filter(
                        product_structure_name=map_prod_class.get(format_string(th_row[0].value),
                                                                  format_string(th_row[0].value))).first()
                    if not bg_class:
                        print("没有该大类")
                        continue

                    md_class = format_string(th_row[1].value)
                    sm_class = format_string(th_row[2].value)

                    pd_cls_obj = ProductsType.objects.filter(
                        Q(project_name="A") & Q(struct_name=bg_class) & Q(prod_type=md_class) & Q(
                            child_type=sm_class)).first()
                    if pd_cls_obj:
                        three_index_obj = Index()
                        three_index_obj.prod_class = pd_cls_obj
                        three_index_obj.first_index = first_index_obj
                        three_index_obj.two_index = two_index_obj
                        three_index_obj.index_explain = th_row[5].value
                        three_index_obj.index_name = th_row[6].value
                        three_index_obj.index_description = th_row[7].value
                        three_index_obj.index_id = str(th_row[8].value)[1:] if th_row[8].value else th_row[8].value
                        three_index_obj.anli_id = str(th_row[9].value)[1:] if th_row[9].value else th_row[9].value
                        three_index_obj.test_type = th_row[10].value
                        three_index_obj.tool = th_row[11].value
                        three_index_obj.remark = th_row[12].value
                        three_index_obj.save()
                    else:
                        print("没有该产品类型")


def index_save_db(request):
    recv_data = request.POST.dict()
    obj_id = recv_data.get("obj_id", "")
    idx_obj = Index.objects.filter(id=int(obj_id)).first()
    idx_obj.index_explain = recv_data.get("index_explain", "")
    idx_obj.index_name = recv_data.get("index_name", "")
    idx_obj.index_description = recv_data.get("index_description", "")
    # 案例ID、指标ID 不做修改
    # idx_obj.index_id = recv_data.get("index_id", "")
    # idx_obj.anli_id = recv_data.get("anli_id", "")
    idx_obj.test_type = recv_data.get("test_type", "")
    idx_obj.tool = recv_data.get("tool", "")
    idx_obj.remark = recv_data.get("remark", "")
    idx_obj.flag = "1"
    idx_obj.save()
    return JsonResponse({"code": "0000"})


def index_add(request):
    recv_data = request.POST.dict()
    pj_name = recv_data.get("pj_name", "")
    pdt_pj_name = "B"

    prod_type = recv_data.get("prod_type", "")
    prod_child_type = recv_data.get("prod_child_type", "")
    first_idx = recv_data.get("first_idx", "")
    two_idx = recv_data.get("two_idx", "")
    if not all([pj_name, prod_child_type, prod_type, first_idx, two_idx]):
        return JsonResponse({"code": "0001"})

    # 存储数据库
    pdt_obj = ProductsType.objects.filter(
        Q(project_name=pdt_pj_name) & Q(prod_type=str(prod_type)) & Q(child_type=str(prod_child_type))).first()
    f_idx_obj = FirstIndex.objects.filter(first_index=first_idx).first()
    t_idx_obj = TwoIndex.objects.filter(Q(first_index=f_idx_obj) & Q(two_index=two_idx)).first()

    if not pdt_obj or not f_idx_obj or not t_idx_obj:
        return JsonResponse({"code": "0002"})
    anli_id_list = [(index_obj.anli_id)[4:] for index_obj in
                    Index.objects.filter(Q(project_name=recv_data.get("pj_name", "")) & Q(two_index=t_idx_obj))]
    if anli_id_list:
        anli_id_new = str(t_idx_obj.index_number) + str(int(max(anli_id_list)) + 1).zfill(4)
    else:
        anli_id_new = str(t_idx_obj.index_number) + str("0001").zfill(4)
    index_new = Index()
    index_new.project_name = recv_data.get("pj_name", "")
    index_new.prod_class = pdt_obj
    index_new.first_index = f_idx_obj
    index_new.two_index = t_idx_obj
    index_new.index_explain = recv_data.get("index_explain", "")
    index_new.index_name = recv_data.get("index_name", "")
    index_new.index_description = recv_data.get("index_description", "")
    index_new.index_id = t_idx_obj.index_number
    index_new.anli_id = anli_id_new
    index_new.test_type = recv_data.get("test_type", "")
    index_new.tool = recv_data.get("tool", "")
    index_new.remark = recv_data.get("remark", "")
    index_new.flag = "2"  # 新增
    index_new.save()
    print("---ppppppppppppppp----------")
    return JsonResponse({"code": "0000"})


# -------

def index_excel_save(pj_name):
    prod_list = [[pd_obj.prod_class.prod_type, pd_obj.prod_class.child_type, pd_obj.first_index.first_index,
                  pd_obj.two_index.two_index,
                  pd_obj.index_explain, pd_obj.index_name, pd_obj.index_description, pd_obj.index_id, pd_obj.anli_id,
                  pd_obj.test_type, pd_obj.tool, pd_obj.remark] for pd_obj in
                 Index.objects.filter(project_name=pj_name)]
    title_list = ["产品类型", "子类", "一级指标", "二级指标", "指标说明", "指标名称", "指标描述", "指标id", "案例编号", "测试类型", "工具", "备注"]
    prod_list.insert(0, title_list)

    file_name = f"index_{pj_name}" + ".xlsx"
    files_mulu = os.getcwd() + "/static/upload_files/index_export"
    file_name_list = os.listdir(files_mulu)
    if file_name in file_name_list:
        os.remove(files_mulu + "/" + file_name)
    # 存储成excel
    wb = Workbook()
    wb.create_sheet("指标信息")
    ws = wb.get_sheet_by_name("指标信息")
    for sv in prod_list:
        ws.append(sv)
    wb.save(files_mulu + "/" + file_name)


def index_export(request):
    pj_name = request.GET.get("pj_name", "")
    index_excel_save(pj_name)
    if pj_name:
        print(pj_name)
        file_name = f"index_{pj_name}" + ".xlsx"
        files_mulu = os.getcwd() + "/static/upload_files/index_export"
        file_list = os.listdir(files_mulu)
        if file_name in file_list:
            wb = openpyxl.load_workbook(files_mulu + "/" + file_name)
            response = HttpResponse(content_type='application/msexcel')
            response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(
                escape_uri_path(f"{pj_name}指标目录.xlsx"))
            wb.save(response)
            return response
        return HttpResponse("文件不存在")
    else:
        return HttpResponse("项目名不存在")


def index_query(request):
    first_idx = request.POST.get("first_idx", "")
    f_obj = FirstIndex.objects.filter(first_index=first_idx).first()
    t_list = [t_obj.two_index for t_obj in TwoIndex.objects.filter(first_index=f_obj)]
    return JsonResponse({"code": "0000", "t_list": t_list})


def index_del(request):
    del_obj_id = request.POST.get("del_obj_id", "")
    Index.objects.filter(id=int(del_obj_id)).delete()
    return JsonResponse({"code": "0000"})

# 保存指标信息到数据库
# def index_info(request):
#     print("plist_save start--------")
#     files_mulu = os.getcwd() + "/static/upload_files/product_list"
#     file_name = "指标数据字典.xlsx"
#     all_list = []
#
#     if 1:
#         file_list = os.listdir(files_mulu)
#         print(file_list)
#         struct_dic = dict()
#         for file_name in file_list:
#             show_mulu = files_mulu + "/" + file_name
#             sheet_obj = openpyxl.load_workbook(show_mulu)
#             print("-------------+++++++")
#             big_class_list = sheet_obj.sheetnames
#             for bg_sheet in big_class_list:
#                 if bg_sheet == "产品目录":
#
#                     bg_table = sheet_obj[bg_sheet]
#                     nrows = list(bg_table.rows)
#                     del nrows[0]
#                     pd_big_class_list = list(set([rw[0].value for rw in nrows if rw[0].value]))
#                     # print(md_class_list)
#                     for big_cls in pd_big_class_list:
#                         struct_dic[big_cls] = {}
#                         # 读取第二列(产品类型)
#                         m_prod_list = [sm_rw[1].value for sm_rw in nrows if sm_rw[0].value and sm_rw[0].value == big_cls]
#                         m_prod_list = list(set(m_prod_list))
#                         # 把子类放到字典里
#                         for m_prd in m_prod_list:
#                             struct_dic[big_cls][m_prd] = {}
#
#
#                             struct_dic[big_cls][m_prd][s_prd] = [
#                                 '%s>%s <span style="display:none;">' % (sm_rw[6].value, sm_rw[5].value) + (
#                                     '%s|%s' % (sm_rw[6].value, sm_rw[5].value) + "|" +
#                                     bg_class + "|" + md_cls + "|" + s_prd if s_prd else "") + '</span>' for sm_rw in
#                                 nrows if
#                                 sm_rw[1].value == s_prd]
#             # print(struct_dic)
#             for big_class in struct_dic:
#                 prod_dic = dict()
#                 prod_dic["title"] = big_class
#                 # 以下数据结构为layui的tree组件所需要的格式
#                 prod_dic["children"] = [{"title": middle_cls if middle_cls else "空", "children": [
#                     {"title": small_cls if small_cls else "空",
#                      "children": [{"title": prod if prod else "空"} for prod in
#                                   struct_dic[big_class][middle_cls][small_cls]]} for
#                     small_cls in struct_dic[big_class][middle_cls]]} for middle_cls in struct_dic[big_class]]
#                 all_list.append(prod_dic)
#
#         return JsonResponse({"code": "0000", "prod_info": all_list})
#
#         # if pd_sheet in sheet_obj.sheetnames:
#         #     pass_table = sheet_obj["产品目录"]
#         #     nrows = pass_table.rows
#         #     all_list = []
#         #     flg = 0
#         #     nrows = list(nrows)
#         #     del nrows[0]
#         #     for r in nrows:
#         #         resp_j = {}
#         #         ll = []
#         #         for i in r:
#         #             ll.append(i.value if i.value else "")
#         #         print(ll)
#         #         break
