# -*- coding:utf-8 -*-
from openpyxl import load_workbook
import os


from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from xuanxing.models import XuanXingRank
# from xuanxing.models import ZhuangXiangRank
# from xuanxing.models import ChushiRank
# from xuanxing.models import AddScoreRank
# from xuanxing.models import MinusScoreRank



def t1():
    wb = Workbook()
    print(wb.sheetnames)  # 提供一个默认名叫Sheet的表，office2016下新建提供默认Sheet1

    a_sheet = wb["Sheet"]
    print("1----------")
    # 直接赋值就可以改工作表的名称
    a_sheet.title = 'Sheet1'
    print("2=--------------")
    # 新建一个工作表，可以指定索引，适当安排其在工作簿中的位置
    wb.create_sheet('Data', index=1)  # 被安排到第二个工作表，index=0就是第一个位置
    print("3-----")

    # 删除某个工作表
    wb.remove(a_sheet)
    print("4-----------")
    # del wb[a_sheet]
    print(wb.sheetnames)
    # ---------------------------------
    sheet = wb.active
    # 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
    b4 = sheet['B4']
    # 分别返回
    print(f'({b4.column}, {b4.row}) is {b4.value}')  # 返回的数字就是int型

    # ------
    # 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
    b4_too = sheet.cell(row=4, column=2)
    b4_too.value = "dadad"
    print(b4_too.value)

    # 获得最大列和最大行
    print(sheet.max_row)
    print(sheet.max_column)
    # ---------------
    # 因为按行，所以返回A1, B1, C1这样的顺序
    for row in sheet.rows:
        for cell in row:
            print(cell.value)
    # A1, A2, A3这样的顺序
    for column in sheet.columns:
        for cell in column:
            print(cell.value)
    # sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引，list(sheet.rows)[2]这样就获取到第三行的tuple对象。
    for cell in list(sheet.rows)[2]:
        print(cell.value)

    # 获得了以A1为左上角，B3为右下角矩形区域的所有单元格
    # 在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值
    for i in range(1, 4):
        for j in range(1, 3):
            print(sheet.cell(row=i, column=j))

    # 还可以像使用切片那样使用。sheet['A1':'B3']返回一个tuple，该元组内部还是元组，由每行的单元格构成一个元组。
    for cell in sheet['A1':'B3']:
        print(cell)

    for row_cell in sheet['A1':'B3']:
        for cell in row_cell:
            print(cell)

    # 根据列的数字返回字母
    print(get_column_letter(2))  # B
    # 根据字母返回列的数字
    print(column_index_from_string('D'))  # 4

    # ===============
    # 以下为写入
    # 直接给单元格赋值就行
    sheet['A1'] = 'good'
    # B9处写入平均值
    # 但是如果是读取的时候需要加上data_only=True这样读到B9返回的就是数字，如果不加这个参数，返回的将是公式本身'=AVERAGE(B2:B8)'
    sheet['B9'] = '=AVERAGE(B2:B8)'

    # append函数
    # 可以一次添加多行数据，从第一行空白行开始（下面都是空白行）写入。

    # 添加一行
    row = [1, 2, 3, 4, 5]
    sheet.append(row)

    # 添加多行
    rows = [
        ['Number', 'data1', 'data2'],
        [2, 40, 30],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 10],
        [6, 25, 5],
        [7, 50, 10],
    ]
    # 由于append函数只能按行写入。如果我们想按列写入呢。append能实现需求么？如果把上面的列表嵌套看作矩阵。只要将矩阵转置就可以了。使用zip()函数可以实现，不过内部的列表变成了元组就是了。都是可迭代对象，不影响。

    zip_rows = list(zip(*rows))

    '''
    上面一行代码的输出结果如下
    [('Number', 2, 3, 4, 5, 6, 7),
     ('data1', 40, 40, 50, 30, 25, 50),
     ('data2', 30, 25, 30, 10, 5, 10)]
    '''
    # 解释下上面的list(zip(*rows))首先*rows将列表打散，相当于填入了若干个参数，zip从某个列表中提取第1个值组合成一个tuple，再从每个列表中提取第2个值组合成一个tuple，一直到最短列表的最后一个值提取完毕后结束，更长列表的之后的值被舍弃，换句话，最后的元组个数是由原来每个参数（可迭代对象）的最短长度决定的。比如现在随便删掉一个值，最短列表长度为2，data2那一列（竖着看）的值全部被舍弃。
    rows = [
        ['Number', 'data1', 'data2'],
        [2, 40],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 10],
        [6, 25, 5],
        [7, 50, 10],
    ]
    # out
    # [('Number', 2, 3, 4, 5, 6, 7), ('data1', 40, 40, 50, 30, 25, 50)]

    # 设置单元格风格--Style
    # 先导入需要的类from openpyxl.styles import Font, colors, Alignment
    # 分别可指定字体相关，颜色，和对齐方式。
    bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)

    sheet['A1'].font = bold_itatic_24_font
    # 对齐方式
    # 也是直接使用cell的属性aligment，这里指定垂直居中和水平居中。除了center，还可以使用right、left等等参数。

    # 设置B1中的数据垂直居中和水平居中
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置行高和列宽 有时候数据太长显示不完，就需要拉长拉高单元格。

    # 第2行行高
    sheet.row_dimensions[2].height = 40
    # C列列宽
    for i in range(1, 10):
        sheet.column_dimensions[get_column_letter(i)].width = 30

    # sheet.column_dimensions['A'].width = 100

    # 合并和拆分单元格
    # 所谓合并单元格，即以合并区域的左上角的那个单元格为基准，覆盖其他单元格使之称为一个大的单元格。
    # 相反，拆分单元格后将这个大单元格的值返回到原来的左上角位置。

    # 合并单元格， 往左上角写入数据即可
    sheet.merge_cells('B1:G1')  # 合并一行中的几个单元格
    sheet.merge_cells('A1:C3')  # 合并一个矩形区域中的单元格

    # 合并后只可以往左上角写入数据，也就是区间中:左边的坐标。
    # 如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。换句话说若合并前不是在左上角写入数据，合并后单元格中不会有数据。
    # 以下是拆分单元格的代码。拆分后，值回到A1位置。
    sheet.unmerge_cells('A1:C3')

    sheet["A2"] = "测试中心科技产品选型处2020年第一二季度人员考核表"
    sheet.merge_cells('A2:H2')  # 合并一个矩形区域中的单元格
    sheet.row_dimensions[2].height = 40
    bold_itatic_16_font = Font(name='等线', size=16, italic=True, color=colors.BLUE, bold=True)

    sheet['A2'].font = bold_itatic_16_font
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

    sheet["A3"] = "工作事项"
    sheet.row_dimensions[3].height = 30
    sheet['A3'].font = Font(size=13, bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')


    sheet.merge_cells("A3:B3")
    sheet.merge_cells("A3:B" + str(3 + 8))
    jizhun_row = 4
    now_row = jizhun_row
    sx_i = 0
    sx_sh_list = ["质量", "效率", "上会效果", "参与比例", "测试环境复杂系数", "测试案例系数", "测试规模系数(详见说明)", "总得分"]
    sx_list = ["质量", "效率", "参与比例", "测试环境复杂系数", "测试案例系数", "测试规模系数(详见说明)", "总得分"]
    member_all = ["戴路", "张勇", "姜炜", "倪海波", "杨博", "王超"]
    all_types = ["选型测试", "专项工作", "处室贡献", "加分项", "扣分项"]
    # 循环全部数据
    # for xx_type in all_types:
    #     start_row = now_row  # 记录开始的行，以便后面合并单元格
    #     sheet["A" + str(now_row)] = xx_type  # 先把大类放进表格

    # sheet["A" + str(jizhun_row)] = ""
    # for shix in shixiang_all:
    #     if shix[1] == "0":
    #         now_row = jizhun_row + sx_i * 8  # 更新当前行
    #         sheet["B" + str(now_row)] = shix[0]
    #
    #     sx_i += 1

    wb.save(r'example2.xlsx')


def t2():
    wb = Workbook()
    sheet = wb["Sheet"]
    # 大标题
    sheet["A2"] = "测试中心科技产品选型处2020年第一二季度人员考核表"
    sheet.merge_cells('A2:H2')  # 合并一个矩形区域中的单元格
    sheet.row_dimensions[2].height = 40
    bold_itatic_16_font = Font(name='等线', size=16, italic=False, color=colors.BLACK, bold=True)
    sheet['A2'].font = bold_itatic_16_font
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
    # 小标题
    sheet["A3"] = "工作事项"
    sheet.row_dimensions[3].height = 30
    sheet['A3'].font = Font(size=13, bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells("A3:B4")  # todo
    # 基本参数
    sx_sh_list = ["质量", "效率", "上会效果", "参与比例", "测试环境复杂系数", "测试案例系数", "测试规模系数(详见说明)", "总得分"]
    sx_list = ["质量", "效率", "参与比例", "测试环境复杂系数", "测试案例系数", "测试规模系数(详见说明)", "总得分"]
    member_all = ["戴路", "张勇", "姜炜", "倪海波", "杨博", "王超"]
    all_types = ["选型测试", "专项工作", "处室贡献", "加分项", "扣分项"]

    jizhun_row = 4
    now_row = jizhun_row
    sx_i = 0

    # 循环全部数据
    for xx_type in all_types:
        start_row = now_row  # 记录开始的行，以便后面合并单元格
        sheet["A" + str(now_row)] = xx_type  # 先把大类放进表格
        # for xx_obj in

    sheet["A" + str(jizhun_row)] = ""
    # for shix in shixiang_all:
    #     if shix[1] == "0":
    #         now_row = jizhun_row + sx_i * 8  # 更新当前行
    #         sheet["B" + str(now_row)] = shix[0]
    #
    #     sx_i += 1


if __name__ == '__main__':
    t1()
