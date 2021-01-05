# -*- coding:utf-8 -*-
import os
import json
import openpyxl
import datetime
from django.shortcuts import render
from django.shortcuts import redirect
from testm.models import Projects
from django.utils.encoding import escape_uri_path
from xuanxing.models import ManageUser
from xuanxing.models import ProjectDaily
from xuanxing.models import Log
from testm.models import Dongtai
from tools.tools import juge_user_pj
from openpyxl import Workbook
from django.http import HttpResponse
from tools.pm_utils_page import PagePaging
from django.http import JsonResponse
from tools.tools import format_string
from django.db.models import Q


def project_dongtai(request):
    d_list = Dongtai.objects.all().values_list('project_id', flat=True)
    dt_list = list(set([dt for dt in d_list]))
    user = ManageUser.objects.filter(id=int(request.session.get("user_id", ""))).first()
    dongtai_list_sorted = []
    if user:
        role_list = [u_obj.rolename for u_obj in user.user_role.all()]
        if "reporter" in role_list or "超级管理员" in role_list:
            dongtai_list = [[dt_obj for dt_obj in Dongtai.objects.filter(project_id=dt_id)] for dt_id in dt_list]
        else:
            project_list = juge_user_pj(request.session.get("user_id", ""))
            dongtai_list = [[dt_obj for dt_obj in Dongtai.objects.filter(project_id=dt_id)] for dt_id in dt_list if
                            Projects.objects.filter(id=dt_id).first() and Projects.objects.filter(
                                id=dt_id).first().project_name in project_list]
        dongtai_list_sorted = [sorted(dongtai, key=lambda x: x.s_time) for dongtai in
                               dongtai_list] if dongtai_list else []
    return render(request, "xuanxing_html/dongtai.html", {"ld": dongtai_list_sorted})


def dt_search(request):
    pro_name = str(request.POST.get('pro_name', ""))
    project_list = juge_user_pj(request.session.get("user_id", ""))
    pj_id_list = [Projects.objects.filter(project_name=pjname).first().id for pjname in project_list if
                  pro_name in pjname]
    p_obj = Dongtai.objects.filter(project_id__in=pj_id_list)

    return render(request, "xuanxing_html/dongtai.html", {"p_obj": p_obj})


def project_export(request):
    if request.method == "POST":
        recv_data = request.POST.dict()
        select_flag = recv_data.get("select_flag", "")
        box_style = recv_data.get("box_style", "")
        if not box_style:
            return JsonResponse({"code": "0004"})
        box_style = json.loads(box_style)
        daochu_list = [daochu_id for daochu_id, daochu_flag in box_style.items() if daochu_flag == "1"]
        # 对全选的进行处理
        if select_flag == "1":
            for daochu_pj in Projects.objects.all():
                if str(daochu_pj.id) not in box_style:
                    daochu_list.append(str(daochu_pj.id))
        d_str = "导出了项目信息"
        opt_usr = ManageUser.objects.filter(id=int(request.session.get("user_id", "0"))).first()
        Log.objects.create(operation=d_str, user_id=opt_usr)
        return JsonResponse({"code": "0000", "check_str": json.dumps(daochu_list)})

    if request.method == "GET":
        file_name = "项目信息.xlsx"
        files_mulu = os.getcwd() + "/static/project"
        file_list = os.listdir(files_mulu)
        if file_name in file_list:
            os.remove(files_mulu + "/" + file_name)
        pj_list = request.GET.get("pj_list", "")
        if not pj_list:
            return HttpResponse("请选择项目")
        pj_list = json.loads(pj_list)
        project_save(file_name, files_mulu, pj_list)
        wb = openpyxl.load_workbook(files_mulu + "/" + file_name)
        print(wb.sheetnames)
        ws = wb["Sheet"]
        wb.remove(ws)
        wb.save(files_mulu + "/" + file_name)
        response = HttpResponse(content_type='application/msexcel')
        response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path("项目信息.xlsx"))
        wb.save(response)
        return response


def project_save(file_name, files_mulu, pj_list):
    meet_status = {"0": "未知", "1": "选型测试(上会)", "2": "选型测试(不上会)", "3": "选型评估"}
    pj_list = [[pj.creation_time.strftime("%Y-%m-%d"), pj.project_name,
                pj.deal_user.all().first().username if pj.deal_user.all() else "",
                ' '.join([mem.username for mem in pj.members.all() if mem]), pj.project_speed.all().first().speed_name,
                "\n".join([vend.vend_name + "|" + vend.product_name for vend in pj.vend_prod.all()]),
                pj.product_subclass.all().first().child_type, meet_status[pj.meeting_status]] for pj in
               Projects.objects.filter(id__in=pj_list)]
    title_list = ["创建时间", "项目名称", "负责人", "参与者", "项目进度", "厂商", "产品子类", "上专家会"]
    pj_list.insert(0, title_list)
    wb = Workbook()
    wb.create_sheet("项目信息")
    ws = wb.get_sheet_by_name("项目信息")
    for pj_info in pj_list:
        ws.append(pj_info)
    wb.save(files_mulu + "/" + file_name)


def project_daily(request):
    if request.method == "GET":
        pj_id = request.GET.get("pj_id", "")
        type_dic = {"1": "日报", "2": "会议"}
        pj_name_list = [{"id": pj.id, "name": pj.project_name} for pj in Projects.objects.all()]
        now_date = str(datetime.datetime.now().strftime("%Y-%m-%d"))
        page_num_list = list(range(3,10))
        pj_dic = {"pj_name_list": pj_name_list, "daily_type": type_dic, "pj_id": pj_id, "now_date": now_date,
                  "page_num_list": page_num_list}
        return render(request, "xuanxing_html/project_daily.html", pj_dic)
    else:
        recv_data = request.POST.dict()
        print(recv_data)
        project_id = recv_data.get("pj_id", "")
        curr_page = int(recv_data.get("current_page", 1))
        pj_all = []
        if project_id:
            pj_all = ProjectDaily.objects.filter(pj_id=int(project_id))
        # 分页开始
        page_msg_num = int(recv_data.get("msg_num", 8))
        pp_obj = PagePaging(len(pj_all), page_msg_num)
        totalpage_num = pp_obj.judge()  # 总页数
        page_obj = pp_obj.re_page(curr_page)  # 返回当前页的分页对象 包含start 和 end
        page_jizhun = 1 if curr_page == 1 else (curr_page - 1) * page_msg_num + 1
        type_dic = {"1": "日报", "2": "会议"}
        data_list = [{"id": pj.id, "pj_name": pj.pj_id.project_name, "name": pj.name, "pj_type": type_dic[str(pj.type)], "file_name": pj.file_name, "status":pj.get_status_display(),"mark": pj.mark, "check_mark": pj.check_mark,"up_time": pj.up_time.strftime("%Y-%m-%d")} for pj in pj_all]

        return JsonResponse({"code": "0000", "data": data_list, "total_page": totalpage_num, "curr_page": curr_page,
                             "page_jizhun": page_jizhun, "url": "daily"})


def project_daily_edit(request):
    if request.method == "GET":
        pd_dic = {}
        recv_data = request.GET.dict()
        pd_id = recv_data.get("pd_id", "")
        if pd_id:
            pd_obj = ProjectDaily.objects.filter(id=int(pd_id)).first()
            if pd_obj:
                daily_type_list = [["1", "日报"], ["2", "会议"]]
                pd_dic = {"pj_id": pd_obj.pj_id.project_name, "pd_name": pd_obj.name,
                          "daily_type_list": daily_type_list,
                          "pd_type": str(pd_obj.type), "markinfo": pd_obj.mark, "pd_file_name": pd_obj.file_name, "pd_id": pd_id}
        return render(request, "xuanxing_html/project_daily_edit.html", pd_dic)
    else:
        recv_data = request.POST.dict()
        pd_id = recv_data.get("pd_id", "")
        pd_name = recv_data.get("pd_name", "")
        pd_type = recv_data.get("pd_type", "")
        pd_mark = recv_data.get("pd_mark", "")
        resp_file = request.FILES.get('file', "")

        if not pd_id or not pd_name:
            return JsonResponse({"code": "0001"})
        check_pd_obj = ProjectDaily.objects.filter(name=pd_name).first()
        if check_pd_obj and str(check_pd_obj.id) != pd_id:
            return JsonResponse({"code": "0002"})

        pd_obj = ProjectDaily.objects.filter(id=int(pd_id)).first()
        if not pd_obj:
            return JsonResponse({"code": "0003"})

        # 存文件
        if resp_file:
            pj_name = pd_obj.pj_id.project_name
            files_mulu = os.getcwd() + "/static/project_daily/"
            ml_list = os.listdir(files_mulu)
            print(ml_list)
            if pj_name not in ml_list:
                os.mkdir(files_mulu + pj_name)
            if resp_file:
                with open('static/project_daily/' + pj_name + "/" + resp_file.name, 'wb') as f:
                    for line in resp_file.chunks():
                        f.write(line)
        # 数据存储到数据库
        c_time=datetime.datetime.now()  #改动s
        pd_obj.name = pd_name
        pd_obj.type = pd_type
        pd_obj.mark = pd_mark
        pd_obj.up_time = c_time    #改动s
        if resp_file:
            pd_obj.file_name = resp_file.name
            pd_obj.status = 2  # 改动
        pd_obj.save()
        return JsonResponse({"code": "0000"})


def project_daily_upload(request):
    recv_data = request.POST.dict()
    resp_file = request.FILES.get('file', "")
    pj_id = recv_data.get("add_pj", "")
    item_name = format_string(recv_data.get("add_item_name", ""))
    if not pj_id or not item_name:
        return JsonResponse({"code": "0001"})
    pj_obj = Projects.objects.filter(id=int(pj_id)).first()
    if not pj_obj:
        return JsonResponse({"code": "0001"})
    pj_name = pj_obj.project_name
    if ProjectDaily.objects.filter(Q(name=item_name) & Q(pj_id=int(pj_id))).first():
        return JsonResponse({"code": "0002"})

    files_mulu = os.getcwd() + "/static/project_daily/"
    ml_list = os.listdir(files_mulu)
    print(ml_list)
    if pj_name not in ml_list:
        os.mkdir(files_mulu + pj_name)
    if resp_file:
        with open('static/project_daily/' + pj_name + "/" + resp_file.name, 'wb') as f:
            for line in resp_file.chunks():
                f.write(line)
    pd_obj = ProjectDaily()
    pd_obj.pj_id = pj_obj
    pd_obj.name = item_name

    pd_obj.type = recv_data.get("add_item_type", "")
    if resp_file:
        print(888)
        pd_obj.file_name = resp_file.name
        pd_obj.status = 2   #改动
    pd_obj.mark = recv_data.get("mark_info", "")
    pd_obj.save()
    return JsonResponse({"code": "0000"})


def project_daily_del(request):
    # 删除项目
    pd_id = request.GET.get("pd_id", "")
    if pd_id:
        obj = ProjectDaily.objects.get(id=int(pd_id))
        if obj:
            obj.delete()
    return redirect("/xuanxing/project_daily/")
